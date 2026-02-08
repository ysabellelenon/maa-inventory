from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404, JsonResponse
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Q, F, DecimalField, Count, Max, Prefetch
from django.db.models.functions import Coalesce
from decimal import Decimal
import json
from .forms import RegistrationForm, LoginForm, SupplierForm, ItemForm, SupplierItemForm, PriceDiscussionForm
from .models import (
    Item, ItemVariation, StockBalance, InventoryLocation,
    Supplier, SupplierCategory, SupplierItem, SupplierOrder, SupplierOrderItem,
    Request, RequestItem, RequestStatusHistory, Branch, Brand, ValidPunchID, UserProfile, Role,
    IntegrationFoodics, ImportJob, SystemSettings, ItemPhoto, PortalToken,
    SupplierPriceDiscussion, BranchPackagingRule, BranchPackagingItem, BranchPackagingRuleItem,
)


def send_invoice_email(supplier_order, request):
    """
    Send purchase order invoice email to supplier with a secure token link to view online
    
    Args:
        supplier_order: SupplierOrder instance
        request: HttpRequest object to build absolute URLs
    """
    from django.core.mail import EmailMultiAlternatives
    from django.conf import settings
    from django.urls import reverse
    
    # Get supplier email
    supplier_email = supplier_order.supplier.email
    if not supplier_email:
        raise ValueError(f"Supplier {supplier_order.supplier.name} has no email address")
    
    # Get requester email (the user who created the order) for CC
    requester_email = None
    if supplier_order.created_by and supplier_order.created_by.email:
        requester_email = supplier_order.created_by.email
    
    # Get the portal token for this order
    portal_token = supplier_order.portal_tokens.first()
    if not portal_token:
        # Fallback to order ID if no token exists (backward compatibility)
        invoice_path = reverse('view_invoice', kwargs={'order_id': supplier_order.id})
        invoice_url = request.build_absolute_uri(invoice_path)
    else:
        # Use secure token-based URL
        invoice_path = reverse('view_invoice_by_token', kwargs={'token': portal_token.token})
        invoice_url = request.build_absolute_uri(invoice_path)
    
    # Calculate order total for email
    order_items = SupplierOrderItem.objects.filter(
        supplier_order=supplier_order
    ).select_related('item')
    
    subtotal = Decimal('0.00')
    item_count = 0
    for order_item in order_items:
        line_total = order_item.qty_ordered * order_item.price_per_unit
        subtotal += line_total
        item_count += 1
    
    # Create email subject
    subject = f'Purchase Order {supplier_order.po_code} - MAA Inventory'
    
    # Plain text body
    text_body = f"""
Dear {supplier_order.supplier.name},

We have created a new purchase order for your review.

Purchase Order: {supplier_order.po_code}
Order Date: {supplier_order.created_at.strftime('%B %d, %Y')}
Items: {item_count}
Total Amount: ${subtotal:,.2f}

Please click the link below to view the full invoice details:

{invoice_url}

If you have any questions, please don't hesitate to contact us.

Best regards,
MAA Inventory Team
"""

    # HTML body
    html_body = f"""
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            line-height: 1.6;
            color: #101828;
            margin: 0;
            padding: 0;
            background-color: #F9FAFB;
        }}
        .container {{
            max-width: 600px;
            margin: 40px auto;
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
        }}
        .header {{
            border-bottom: 3px solid #D9BD7D;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }}
        h1 {{
            color: #D9BD7D;
            font-size: 24px;
            margin: 0 0 10px 0;
        }}
        .po-number {{
            font-size: 18px;
            color: #475467;
            margin: 0;
        }}
        .details {{
            background: #F9FAFB;
            padding: 20px;
            border-radius: 8px;
            margin: 20px 0;
        }}
        .details-row {{
            display: flex;
            justify-content: space-between;
            padding: 8px 0;
            border-bottom: 1px solid #E5E7EB;
        }}
        .details-row:last-child {{
            border-bottom: none;
        }}
        .label {{
            color: #6B7280;
            font-weight: 500;
        }}
        .value {{
            color: #101828;
            font-weight: 600;
        }}
        .button {{
            display: inline-block;
            background-color: #D9BD7D;
            color: white;
            text-decoration: none;
            padding: 14px 28px;
            border-radius: 8px;
            font-weight: 600;
            margin: 20px 0;
            text-align: center;
        }}
        .button:hover {{
            background-color: #C7AB6B;
        }}
        .footer {{
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #E5E7EB;
            color: #6B7280;
            font-size: 14px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Purchase Order Created</h1>
            <p class="po-number">{supplier_order.po_code}</p>
        </div>
        
        <p>Dear {supplier_order.supplier.name},</p>
        
        <p>We have created a new purchase order for your review.</p>
        
        <div class="details">
            <div class="details-row">
                <span class="label">Order Date:</span>
                <span class="value">{supplier_order.created_at.strftime('%B %d, %Y')}</span>
            </div>
            <div class="details-row">
                <span class="label">Items:</span>
                <span class="value">{item_count}</span>
            </div>
            <div class="details-row">
                <span class="label">Total Amount:</span>
                <span class="value">${subtotal:,.2f}</span>
            </div>
        </div>
        
        <div style="text-align: center;">
            <a href="{invoice_url}" class="button">View Invoice</a>
        </div>
        
        <div class="footer">
            <p>If you have any questions, please don't hesitate to contact us.</p>
            <p>Best regards,<br>MAA Inventory Team</p>
        </div>
    </div>
</body>
</html>
"""
    
    # Create email message
    email_recipients = [supplier_email]
    email_cc = []
    
    # Add requester to CC if they have an email
    if requester_email:
        email_cc.append(requester_email)
    
    email = EmailMultiAlternatives(
        subject=subject,
        body=text_body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=email_recipients,
        cc=email_cc if email_cc else None,
    )
    
    # Attach HTML version
    email.attach_alternative(html_body, "text/html")
    
    # Send email
    email.send(fail_silently=False)


def send_receiving_note_email(cancelled_order, new_order, note, request):
    """
    Send single email when order is cancelled and recreated
    Includes cancellation notice, warehouse note, and link to new invoice
    
    Args:
        cancelled_order: The cancelled SupplierOrder instance
        new_order: The new SupplierOrder instance
        note: The note text from warehouse staff
        request: HttpRequest object to build absolute URLs
    """
    from django.core.mail import EmailMultiAlternatives
    from django.conf import settings
    from django.utils import timezone
    from django.urls import reverse
    
    # Get supplier email
    supplier_email = cancelled_order.supplier.email
    if not supplier_email:
        raise ValueError(f"Supplier {cancelled_order.supplier.name} has no email address")
    
    # Get requester email (the user who created the order) for CC
    requester_email = None
    if cancelled_order.created_by and cancelled_order.created_by.email:
        requester_email = cancelled_order.created_by.email
    
    # Get warehouse staff name who added the note
    warehouse_staff_name = request.user.get_full_name() or request.user.username
    if hasattr(request.user, 'profile') and request.user.profile.full_name:
        warehouse_staff_name = request.user.profile.full_name
    
    # Get cancelled invoice URL
    cancelled_portal_token = cancelled_order.portal_tokens.first()
    if cancelled_portal_token:
        cancelled_invoice_path = reverse('view_invoice_by_token', kwargs={'token': cancelled_portal_token.token})
        cancelled_invoice_url = request.build_absolute_uri(cancelled_invoice_path)
    else:
        cancelled_invoice_path = reverse('view_invoice', kwargs={'order_id': cancelled_order.id})
        cancelled_invoice_url = request.build_absolute_uri(cancelled_invoice_path)
    
    # Get new invoice URL
    new_portal_token = new_order.portal_tokens.first()
    if new_portal_token:
        new_invoice_path = reverse('view_invoice_by_token', kwargs={'token': new_portal_token.token})
        new_invoice_url = request.build_absolute_uri(new_invoice_path)
    else:
        new_invoice_path = reverse('view_invoice', kwargs={'order_id': new_order.id})
        new_invoice_url = request.build_absolute_uri(new_invoice_path)
    
    # Calculate order total for email
    from decimal import Decimal
    order_items = new_order.items.all()
    subtotal = Decimal('0.00')
    item_count = 0
    for order_item in order_items:
        line_total = order_item.qty_ordered * order_item.price_per_unit
        subtotal += line_total
        item_count += 1
    
    # Create email subject
    subject = f'Purchase Order {new_order.po_code} - Order Cancelled & Recreated - MAA Inventory'
    
    # Plain text body
    text_body = f"""
Dear {cancelled_order.supplier.name},

We have cancelled Purchase Order {cancelled_order.po_code} and created a new order {new_order.po_code} due to the following warehouse note:

{note}

IMPORTANT: Please review and confirm the new purchase order below.

New Purchase Order: {new_order.po_code}
Order Date: {cancelled_order.created_at.strftime('%B %d, %Y')}
Items: {item_count}
Total Amount: OMR {subtotal:,.2f}
Noted By: {warehouse_staff_name}
Date: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}

Please review both invoices:
- Cancelled Invoice: {cancelled_invoice_url}
- New Invoice: {new_invoice_url}

Note: The previous order {cancelled_order.po_code} has been cancelled.

Please contact us if you have any questions or concerns.

Best regards,
MAA Inventory Team
"""
    
    # HTML body
    html_body = f"""
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            line-height: 1.6;
            color: #101828;
            margin: 0;
            padding: 0;
            background-color: #F9FAFB;
        }}
        .container {{
            max-width: 600px;
            margin: 40px auto;
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
        }}
        .header {{
            border-bottom: 3px solid #F59E0B;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }}
        h1 {{
            color: #F59E0B;
            font-size: 24px;
            margin: 0 0 10px 0;
        }}
        .po-number {{
            font-size: 18px;
            color: #475467;
            margin: 0;
        }}
        .note-box {{
            background: #FEF3C7;
            border-left: 4px solid #F59E0B;
            padding: 20px;
            border-radius: 8px;
            margin: 20px 0;
        }}
        .note-box p {{
            margin: 0;
            color: #92400E;
            font-weight: 500;
        }}
        .details {{
            background: #F9FAFB;
            padding: 20px;
            border-radius: 8px;
            margin: 20px 0;
        }}
        .details-row {{
            display: flex;
            justify-content: space-between;
            padding: 8px 0;
            border-bottom: 1px solid #E5E7EB;
        }}
        .details-row:last-child {{
            border-bottom: none;
        }}
        .label {{
            color: #6B7280;
            font-weight: 500;
        }}
        .value {{
            color: #101828;
            font-weight: 600;
        }}
        .button {{
            display: inline-block;
            background-color: #D9BD7D;
            color: white;
            text-decoration: none;
            padding: 14px 28px;
            border-radius: 8px;
            font-weight: 600;
            margin: 10px 5px;
            text-align: center;
        }}
        .button:hover {{
            background-color: #C7AB6B;
        }}
        .footer {{
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #E5E7EB;
            color: #6B7280;
            font-size: 14px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Order Cancelled & Recreated</h1>
            <p class="po-number">Old PO: {cancelled_order.po_code} → New PO: {new_order.po_code}</p>
        </div>
        
        <p>Dear {cancelled_order.supplier.name},</p>
        
        <p>We have cancelled Purchase Order <strong>{cancelled_order.po_code}</strong> and created a new order <strong>{new_order.po_code}</strong> due to the following warehouse note:</p>
        
        <div class="note-box">
            <p>{note}</p>
        </div>
        
        <p style="font-weight: 600; color: #DC2626; margin: 20px 0;">IMPORTANT: Please review and confirm the new purchase order below.</p>
        
        <div class="details">
            <div class="details-row">
                <span class="label">New Purchase Order:</span>
                <span class="value">{new_order.po_code}</span>
            </div>
            <div class="details-row">
                <span class="label">Order Date:</span>
                <span class="value">{cancelled_order.created_at.strftime('%B %d, %Y')}</span>
            </div>
            <div class="details-row">
                <span class="label">Items:</span>
                <span class="value">{item_count}</span>
            </div>
            <div class="details-row">
                <span class="label">Total Amount:</span>
                <span class="value">OMR {subtotal:,.2f}</span>
            </div>
            <div class="details-row">
                <span class="label">Noted By:</span>
                <span class="value">{warehouse_staff_name}</span>
            </div>
            <div class="details-row">
                <span class="label">Date:</span>
                <span class="value">{timezone.now().strftime('%B %d, %Y at %I:%M %p')}</span>
            </div>
            <div class="details-row" style="border-top: 2px solid #E5E7EB; margin-top: 8px; padding-top: 12px;">
                <span class="label" style="color: #DC2626;">Cancelled PO:</span>
                <span class="value" style="color: #DC2626; text-decoration: line-through;">{cancelled_order.po_code}</span>
            </div>
        </div>
        
        <div style="text-align: center; margin: 30px 0;">
            <a href="{cancelled_invoice_url}" class="button" style="background-color: #DC2626; font-size: 16px; padding: 16px 32px; margin-right: 12px;">View Cancelled Invoice</a>
            <a href="{new_invoice_url}" class="button" style="background-color: #D9BD7D; font-size: 16px; padding: 16px 32px;">View New Invoice</a>
        </div>
        
        <div class="footer">
            <p><strong>Note:</strong> The previous order {cancelled_order.po_code} has been cancelled.</p>
            <p>Please review both invoices and contact us if you have any questions or concerns.</p>
            <p>Best regards,<br>MAA Inventory Team</p>
        </div>
    </div>
</body>
</html>
"""
    
    # Create email message
    email_recipients = [supplier_email]
    email_cc = []
    
    # Add requester to CC if they have an email
    if requester_email:
        email_cc.append(requester_email)
    
    email = EmailMultiAlternatives(
        subject=subject,
        body=text_body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=email_recipients,
        cc=email_cc if email_cc else None,
    )
    
    # Attach HTML version
    email.attach_alternative(html_body, "text/html")
    
    # Send email
    email.send(fail_silently=False)


def user_login(request):
    """Handle user login using email instead of username"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('email')
            password = form.cleaned_data.get('password')
            
            # Find user by email
            try:
                user = User.objects.get(email=email)
                # Authenticate using username (Django's authenticate requires username)
                user = authenticate(username=user.username, password=password)
                if user is not None:
                    login(request, user)
                    # Get full name from profile if available
                    user_profile = getattr(user, 'profile', None)
                    display_name = user_profile.full_name if user_profile and user_profile.full_name else user.username
                    messages.success(request, f'Welcome back, {display_name}!')
                    next_url = request.GET.get('next', 'dashboard')
                    return redirect(next_url)
                else:
                    messages.error(request, 'Invalid email or password.')
            except User.DoesNotExist:
                messages.error(request, 'Invalid email or password.')
        else:
            messages.error(request, 'Invalid email or password.')
    else:
        form = LoginForm()
    
    return render(request, 'maainventory/login.html', {'form': form})


def user_logout(request):
    """Handle user logout"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


@login_required
def dashboard(request):
    """Render dashboard with dynamic alerts and system information."""
    from django.contrib.auth.models import User
    from django.db.models import Sum, Q, F, DecimalField
    from django.db.models.functions import Coalesce
    from datetime import datetime, timedelta
    from decimal import Decimal
    from .models import (
        UserProfile, Role, Brand, Branch, BranchUser,
        Item, Supplier, Request, SupplierOrder,
        IntegrationFoodics, ImportJob, SystemSettings,
        StockBalance, InventoryLocation, ItemConsumptionDaily, RequestItem
    )
    
    # Get user role to determine dashboard type
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_it = user_role and ('IT' in user_role or user_role == 'IT')
    is_procurement = user_role and 'Procurement' in user_role
    
    # IT users can see both IT dashboard and Procurement dashboard
    # Procurement users only see Procurement dashboard
    show_procurement_alerts = is_procurement or is_it
    
    # ========================================================================
    # ALERT 1: Low Stock Items (based on MIN - min_stock_qty)
    # ========================================================================
    low_stock_items = []
    low_stock_item_ids = set()  # Track unique items to avoid duplicates
    
    # Get all active items
    active_items = Item.objects.filter(is_active=True).select_related('brand')
    
    for item in active_items:
        # Calculate total stock across all warehouse locations for this item
        total_stock = StockBalance.objects.filter(
            item=item,
            location__type='WAREHOUSE'
        ).aggregate(
            total=Coalesce(Sum('qty_on_hand'), Decimal('0'), output_field=DecimalField())
        )['total'] or Decimal('0')
        
        # Check if total stock is below min_stock_qty
        if total_stock < item.min_stock_qty:
            # Only add if we haven't already added this item
            if item.id not in low_stock_item_ids:
                low_stock_item_ids.add(item.id)
                low_stock_items.append({
                    'item_code': item.item_code,
                    'item_name': item.name,
                    'location': 'All Warehouses',
                    'qty_on_hand': float(total_stock),
                    'min_stock_qty': float(item.min_stock_qty),
                    'base_unit': str(item.base_unit) if item.base_unit else '',
                    'shortage': float(item.min_stock_qty - total_stock)
                })
    
    # Sort by shortage amount (most critical first)
    low_stock_items.sort(key=lambda x: x['shortage'], reverse=True)
    low_stock_count = len(low_stock_items)
    
    # ========================================================================
    # ALERT 2: Pending Requests Waiting for Review
    # ========================================================================
    pending_requests = Request.objects.filter(
        status='Pending'
    ).select_related('branch', 'requested_by', 'branch__brand').order_by('-created_at')[:10]
    
    pending_requests_count = Request.objects.filter(status='Pending').count()
    
    # ========================================================================
    # ALERT 3: Items Held at Supplier (Supplier Stock)
    # ========================================================================
    from .models import SupplierStock
    
    supplier_hold_items = []
    supplier_stock_items = SupplierStock.objects.select_related(
        'supplier', 'item', 'variation'
    ).filter(quantity__gt=0).order_by('-confirmed_at')
    
    for stock in supplier_stock_items:
        supplier_hold_items.append({
            'item_code': stock.item.item_code,
            'item_name': stock.item.name,
            'variation': stock.variation.variation_name if stock.variation else None,
            'supplier': stock.supplier.name,
            'location': f"{stock.supplier.name} Stock",
            'qty_on_hand': float(stock.quantity),
            'base_unit': stock.item.base_unit,
            'min_stock_qty': float(stock.item.min_stock_qty),
            'is_low_stock': stock.quantity < stock.item.min_stock_qty,
            'confirmed_at': stock.confirmed_at.strftime("%B %d, %Y") if stock.confirmed_at else "—",
        })
    
    supplier_hold_count = len(supplier_hold_items)
    
    # ========================================================================
    # ALERT 4: Items Requiring Ordering Soon (based on consumption trends)
    # ========================================================================
    items_need_ordering = []
    
    # Get items with consumption data from last 30 days
    thirty_days_ago = datetime.now().date() - timedelta(days=30)
    
    # Calculate consumption per item from Foodics
    consumption_data = ItemConsumptionDaily.objects.filter(
        date__gte=thirty_days_ago
    ).values('item', 'variation').annotate(
        total_consumed=Sum('qty_consumed')
    )
    
    # Calculate requested quantities from pending/warehouse-processing/ready/in-process requests
    request_data = RequestItem.objects.filter(
        request__status__in=['Pending', 'WarehouseProcessing', 'ReadyForDelivery', 'InProcess']
    ).values('item', 'variation').annotate(
        total_requested=Sum('qty_requested')
    )
    
    # Get all active items
    active_items = Item.objects.filter(is_active=True).select_related('brand')
    
    for item in active_items:
        # Get current stock in warehouse
        warehouse_stock = StockBalance.objects.filter(
            item=item,
            location__type='WAREHOUSE'
        ).aggregate(
            total_stock=Coalesce(Sum('qty_on_hand'), Decimal('0'), output_field=DecimalField())
        )['total_stock'] or Decimal('0')
        
        # Get consumption from Foodics (last 30 days)
        item_consumption = consumption_data.filter(item=item).order_by('item', 'variation').first()
        daily_avg_consumption = Decimal('0')
        if item_consumption:
            daily_avg_consumption = item_consumption['total_consumed'] / Decimal('30')
        
        # Get pending requests
        item_requests = request_data.filter(item=item).order_by('item', 'variation').first()
        pending_qty = item_requests['total_requested'] if item_requests else Decimal('0')
        
        # Calculate days until stockout (if consumption continues at current rate)
        days_until_stockout = None
        if daily_avg_consumption > 0:
            days_until_stockout = float(warehouse_stock / daily_avg_consumption) if daily_avg_consumption > 0 else None
        
        # Flag items that need ordering soon (less than 14 days of stock or high pending requests)
        if (days_until_stockout and days_until_stockout < 14) or (pending_qty > warehouse_stock * Decimal('0.5')):
            items_need_ordering.append({
                'item_code': item.item_code,
                'item_name': item.name,
                'current_stock': float(warehouse_stock),
                'min_stock_qty': float(item.min_stock_qty),
                'daily_avg_consumption': float(daily_avg_consumption),
                'pending_requests': float(pending_qty),
                'days_until_stockout': days_until_stockout,
                'base_unit': item.base_unit,
                'urgency': 'high' if (days_until_stockout and days_until_stockout < 7) else 'medium'
            })
    
    # Sort by urgency and days until stockout
    items_need_ordering.sort(key=lambda x: (x['urgency'] == 'high', x.get('days_until_stockout') or 999))
    items_need_ordering_count = len(items_need_ordering)
    
    # ========================================================================
    # Statistics (for both IT and Procurement)
    # ========================================================================
    total_items = Item.objects.filter(is_active=True).count()
    total_suppliers = Supplier.objects.filter(is_active=True).count()
    total_branches = Branch.objects.filter(is_active=True).count()
    active_orders = SupplierOrder.objects.exclude(status__in=['Cancelled', 'Received']).count()
    
    # Additional IT stats
    if is_it:
        total_users = User.objects.count()
        active_users = User.objects.filter(is_active=True).count()
        total_roles = Role.objects.count()
        total_brands = Brand.objects.count()
        import_jobs_total = ImportJob.objects.count()
        import_jobs_pending = ImportJob.objects.filter(status='Pending').count()
        import_jobs_failed = ImportJob.objects.filter(status='Failed').count()
        
        foodics_integration = IntegrationFoodics.objects.first()
        foodics_enabled = foodics_integration.is_enabled if foodics_integration else False
        
        stats = [
            {"label": "Total Users", "key": "stat-users", "value": total_users, "note": f"{active_users} active"},
            {"label": "Roles", "key": "stat-roles", "value": total_roles, "note": "Defined"},
            {"label": "Brands", "key": "stat-brands", "value": total_brands, "note": "Active"},
            {"label": "Branches", "key": "stat-branches", "value": total_branches, "note": "Active"},
            {"label": "Items", "key": "stat-items", "value": total_items, "note": "In system"},
            {"label": "Suppliers", "key": "stat-suppliers", "value": total_suppliers, "note": "Active"},
        ]
    else:
        # Show all statistics cards (some may show 0 if no data yet)
        stats = [
            {"label": "Low Stock", "key": "stat-low-stock", "value": low_stock_count, "note": "Items"},
            {"label": "Pending Requests", "key": "stat-pending", "value": pending_requests_count, "note": "Awaiting review"},
            {"label": "Supplier Stock", "key": "stat-supplier-stock", "value": supplier_hold_count, "note": "Items"},
            {"label": "Need Ordering", "key": "stat-need-ordering", "value": items_need_ordering_count, "note": "Items"},
            {"label": "Active Orders", "key": "stat-active-orders", "value": active_orders, "note": "POs"},
            {"label": "Total Items", "key": "stat-total-items", "value": total_items, "note": "In system"},
    ]

    context = {
        "stats": stats,
        "low_stock_items": low_stock_items[:10],  # Top 10 most critical
        "low_stock_count": low_stock_count,
        "pending_requests": pending_requests,
        "pending_requests_count": pending_requests_count,
        "supplier_hold_items": supplier_hold_items[:5],  # Top 5 preview
        "supplier_hold_count": supplier_hold_count,
        "items_need_ordering": items_need_ordering[:10],  # Top 10 most urgent
        "items_need_ordering_count": items_need_ordering_count,
        "is_it": is_it,
        "is_procurement": is_procurement,
        "show_procurement_alerts": show_procurement_alerts,  # IT and Procurement can see alerts
        "active_orders": active_orders,
    }
    
    # Add IT-specific data
    if is_it:
        recent_users = UserProfile.objects.select_related('user', 'role').order_by('-created_at')[:5]
        recent_imports = ImportJob.objects.select_related('uploaded_by').order_by('-created_at')[:5]
        foodics_integration = IntegrationFoodics.objects.first()
        foodics_last_sync = foodics_integration.last_sync_at if foodics_integration and foodics_integration.last_sync_at else None
        
        role_distribution = []
        for role in Role.objects.all():
            count = UserProfile.objects.filter(role=role).count()
            if count > 0:
                role_distribution.append({"name": role.name, "count": count})
        
        context.update({
            "recent_users": recent_users,
            "recent_imports": recent_imports,
            "foodics_enabled": foodics_integration.is_enabled if foodics_integration else False,
            "foodics_last_sync": foodics_last_sync,
            "import_jobs_total": import_jobs_total,
            "import_jobs_pending": import_jobs_pending,
            "import_jobs_failed": import_jobs_failed,
            "role_distribution": role_distribution,
        })

    return render(request, "maainventory/dashboard.html", context)

@login_required
def inventory(request):
    """Render inventory list page with dynamic data from database."""
    from django.core.paginator import Paginator

    # Supplier categories for the Category dropdown filter
    categories = SupplierCategory.objects.filter(is_active=True).order_by('name')
    category_id = request.GET.get('category')

    # Get all active items; prefetch supplier_items for category (from supplier's category), photos for thumbnails
    items_queryset = Item.objects.filter(is_active=True).select_related('brand').prefetch_related(
        'variations',
        'photos',
        Prefetch('supplier_items', queryset=SupplierItem.objects.select_related('supplier__category')),
    )
    if category_id:
        items_queryset = items_queryset.filter(
            supplier_items__supplier__category_id=category_id
        ).distinct()

    # Get warehouse locations
    warehouse_locations = InventoryLocation.objects.filter(type='WAREHOUSE')

    # Build items list with stock information
    items = []
    for item in items_queryset:
        # Category from first supplier's category (supplier category)
        first_si = next(iter(item.supplier_items.all()), None)
        category_name = None
        if first_si and first_si.supplier and first_si.supplier.category:
            category_name = first_si.supplier.category.name

        # Calculate total stock across all warehouse locations
        total_stock = StockBalance.objects.filter(
            item=item,
            location__type='WAREHOUSE'
        ).aggregate(
            total=Coalesce(Sum('qty_on_hand'), Decimal('0'), output_field=DecimalField())
        )['total'] or Decimal('0')

        # Determine status based on remaining quantity vs min_stock_qty
        status = "LOW" if total_stock < item.min_stock_qty else "GOOD"

        # Get price from Item's price_per_unit field
        price = "—"
        if item.price_per_unit:
            price = f"{item.price_per_unit:.2f}"

        # Image for item name column: photo_url, or first ItemPhoto, or None (template uses fallback)
        image_url = item.photo_url
        if not image_url and item.photos.exists():
            first_photo = item.photos.first()
            if first_photo and first_photo.photo:
                image_url = request.build_absolute_uri(first_photo.photo.url)

        items.append({
            "code": item.item_code,
            "name": item.name,
            "category": category_name,
            "min_stock_qty": f"{item.min_stock_qty:,.0f}",
            "status": status,
            "base_unit": item.base_unit,
            "price": price,
            "qty": f"{total_stock:,.0f}",
            "remaining": f"{total_stock:,.0f}",
            "remaining_qty": total_stock,
            "id": item.id,
            "image": image_url,
        })

    # Paginate items
    paginator = Paginator(items, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        "items": page_obj,
        "page_obj": page_obj,
        "categories": categories,
        "current_category": category_id,
    }

    return render(request, "maainventory/inventory.html", context)


@login_required
def delete_item(request, code):
    """Delete an item (soft delete by setting is_active=False)"""
    item = get_object_or_404(Item, item_code=code)
    
    # Check if user has permission (Procurement Manager or IT)
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_procurement = user_role and 'Procurement' in user_role
    is_it = user_role and ('IT' in user_role or user_role == 'IT')
    
    if not (is_procurement or is_it):
        messages.error(request, 'You do not have permission to delete items.')
        return redirect('inventory')
    
    # Soft delete by setting is_active=False
    item_name = item.name
    item.is_active = False
    item.save()
    messages.success(request, f'Item "{item_name}" has been deleted successfully.')
    return redirect('inventory')


@login_required
def edit_item(request, code):
    """Render edit page for a single inventory item with dynamic data and handle form submission."""
    # Get item by code - try Item.item_code first, then SupplierItem.item_code
    # Allow editing inactive items (similar to suppliers)
    item = None
    
    # First try to find by Item.item_code (without is_active filter to allow editing inactive items)
    try:
        item = Item.objects.get(item_code=code)
    except Item.DoesNotExist:
        # Try to find via SupplierItem.item_code
        supplier_item = SupplierItem.objects.filter(
            item_code=code
        ).select_related('item').first()
        
        if supplier_item and supplier_item.item:
            item = supplier_item.item
    
    if not item:
        raise Http404("Item not found")
    
    # Handle form submission
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            saved_item = form.save(commit=False)
            # Preserve is_active status - don't let form change it
            saved_item.is_active = item.is_active
            saved_item.save()
            
            # Update branches for the Item (not SupplierItem)
            branch_ids = request.POST.getlist('branches')
            branch_ids = [int(bid) for bid in branch_ids if bid.isdigit()]
            item.branches.set(branch_ids)
            
            # Handle photo uploads (maximum 5 total)
            photos = request.FILES.getlist('photos')
            if photos:
                # Check existing photos count
                existing_count = ItemPhoto.objects.filter(item=item).count()
                remaining_slots = 5 - existing_count
                
                if remaining_slots > 0:
                    photos_to_save = photos[:remaining_slots]  # Only save up to the limit
                    saved_count = 0
                    for idx, photo in enumerate(photos_to_save):
                        try:
                            ItemPhoto.objects.create(
                                item=item,
                                photo=photo,
                                order=existing_count + idx
                            )
                            saved_count += 1
                        except Exception as e:
                            messages.error(request, f'Error saving photo {photo.name}: {str(e)}')
                    
                    if saved_count > 0:
                        messages.success(request, f'{saved_count} photo(s) uploaded successfully.')
                    
                    if len(photos) > remaining_slots:
                        messages.warning(request, f'Only {remaining_slots} photo(s) were saved. Maximum 5 photos per item.')
                else:
                    messages.warning(request, 'Maximum 5 photos already exist for this item. Please remove some photos before adding more.')
            
            messages.success(request, f'Item "{item.name}" updated successfully.')
            return redirect('edit_item', code=code)
        else:
            error_fields = [form.fields[name].label for name in form.errors]
            messages.error(
                request,
                'Please fix the following required fields: ' + ', '.join(error_fields) + '.'
            )
    else:
        form = ItemForm(instance=item)
    
    # Get total stock across all warehouse locations
    total_stock = StockBalance.objects.filter(
        item=item,
        location__type='WAREHOUSE'
    ).aggregate(
        total=Coalesce(Sum('qty_on_hand'), Decimal('0'), output_field=DecimalField())
    )['total'] or Decimal('0')
    
    # Determine status
    status = "LOW" if total_stock < item.min_stock_qty else "GOOD"
    
    # Get primary supplier
    primary_supplier_item = SupplierItem.objects.filter(
        item=item,
        is_active=True
    ).select_related('supplier').first()
    
    supplier_name = primary_supplier_item.supplier.name if primary_supplier_item else "—"
    
    # Build item data for template
    selected_item = {
        "code": item.item_code,
        "name": item.name,
        "desc": item.description or "",
        "initial_qty": f"{total_stock:,.0f}",
        "remaining_qty": f"{total_stock:,.0f}",
        "min_qty": f"{item.min_stock_qty:,.0f}",
                "status": status,
        "category": item.brand.name,
        "base_unit": item.base_unit,
        "supplier": supplier_name,
    }
    
    # Get all branches grouped by brand for "Active Branches" section
    all_branches = Branch.objects.filter(is_active=True).select_related('brand').order_by('brand__name', 'name')
    
    # Get branches that currently use this item (from Item.branches)
    branches_using_item_ids = set(
        item.branches.values_list('id', flat=True)
    )
    
    # Group branches by brand
    branch_groups = {}
    for branch in all_branches:
        brand_name = branch.brand.name.upper()
        
        if brand_name not in branch_groups:
            branch_groups[brand_name] = []
        
        branch_groups[brand_name].append({
            'id': branch.id,
            'name': branch.name,
            'brand': branch.brand.name,
            'is_using': branch.id in branches_using_item_ids,
        })
    
    # Get branches that use this item (for display in "Branches Using This Item" section)
    branches_using_item = Branch.objects.filter(
        id__in=branches_using_item_ids
    ).select_related('brand').order_by('brand__name', 'name')
    
    # Get all photos for this item
    item_photos = ItemPhoto.objects.filter(item=item).order_by('order', 'uploaded_at')

    context = {
        "item": selected_item,
        "form": form,
        "branch_groups": branch_groups,
        "branches_using_item": branches_using_item,
        "item_photos": item_photos,
    }

    return render(request, "maainventory/edit_item.html", context)


@login_required
def delete_item_photo(request, photo_id):
    """Delete an item photo"""
    try:
        photo = get_object_or_404(ItemPhoto, id=photo_id)
        item_code = photo.item.item_code
        
        # Check if user has permission (Procurement Manager or IT)
        user_profile = getattr(request.user, 'profile', None)
        user_role = user_profile.role.name if user_profile and user_profile.role else None
        is_procurement = user_role and 'Procurement' in user_role
        is_it = user_role and ('IT' in user_role or user_role == 'IT')
        
        if not (is_procurement or is_it):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
            messages.error(request, 'You do not have permission to delete photos.')
            return redirect('edit_item', code=item_code)
        
        # Delete the photo file from storage
        if photo.photo:
            photo.photo.delete(save=False)
        
        # Delete the photo record
        photo.delete()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Photo deleted successfully'})
        
        messages.success(request, 'Photo deleted successfully.')
        return redirect('edit_item', code=item_code)
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
        messages.error(request, f'Error deleting photo: {str(e)}')
        return redirect('edit_item', code=item_code) if 'item_code' in locals() else redirect('inventory')

@login_required
def requests(request):
    """Render requests page with dynamic data from database. Branch users see only requests for their branch(es)."""
    from django.core.paginator import Paginator
    from django.utils import timezone
    from datetime import timedelta
    from .context_processors import get_branch_user_info

    # Get all requests, ordered by most recent
    requests_queryset = Request.objects.select_related(
        'branch', 'branch__brand', 'requested_by', 'approved_by'
    ).prefetch_related('items__item', 'items__item__photos').order_by('-created_at')

    # Branch managers see only requests for their assigned branch(es); must have assignments
    is_branch_user, user_branch_ids = get_branch_user_info(request.user)
    if is_branch_user:
        if user_branch_ids:
            requests_queryset = requests_queryset.filter(branch_id__in=user_branch_ids)
        else:
            requests_queryset = requests_queryset.none()  # No assignments = see nothing

    # Branch dropdown filter: options are all active branches from the database
    branch_id = request.GET.get('branch')
    branches = Branch.objects.filter(is_active=True).order_by('name')
    if branch_id:
        requests_queryset = requests_queryset.filter(branch_id=branch_id)

    # Tab counts (actual totals): New = requested today or within last 2 days; Pending = status Pending
    now = timezone.now()
    new_cutoff = now - timedelta(days=3)  # today + yesterday + day before = last 3 days
    tab_counts = {
        'all': requests_queryset.count(),
        'new': requests_queryset.filter(date_of_order__gte=new_cutoff).count(),
        'pending': requests_queryset.filter(status=Request.StatusType.PENDING).count(),
        'in-process': 0,
        'delivered': 0,
        'completed': 0,
        'rejected': 0,
    }
    
    # Build requests list for template
    requests_list = []
    for req in requests_queryset:
        # Get first item from request for display
        first_item = req.items.first()
        item_name = first_item.item.name if first_item else "Multiple items"
        request_date = req.date_of_order or req.created_at
        is_new = request_date >= new_cutoff

        # Image for first item: photo_url or first ItemPhoto
        image_url = None
        if first_item:
            item_obj = first_item.item
            image_url = item_obj.photo_url
            if not image_url and item_obj.photos.exists():
                first_photo = item_obj.photos.first()
                if first_photo and first_photo.photo:
                    image_url = request.build_absolute_uri(first_photo.photo.url)
        
        requests_list.append({
            "code": req.request_code,
            "requestor": req.requested_by.profile.full_name if hasattr(req.requested_by, 'profile') and req.requested_by.profile.full_name else (req.requested_by.get_full_name() or req.requested_by.username),
            "branch": req.branch.name,
            "name": item_name,
            "requested_date": req.date_of_order.strftime("%m/%d/%Y") if req.date_of_order else req.created_at.strftime("%m/%d/%Y"),
            "status": req.get_status_display(),
            "id": req.id,  # For detail links
            "is_new": is_new,
            "image": image_url,
        })

    # Paginate requests
    paginator = Paginator(requests_list, 10)  # 10 items per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        "items": page_obj,
        "page_obj": page_obj,
        "tab_counts": tab_counts,
        "branches": branches,
        "current_branch": branch_id,
    }
    return render(request, "maainventory/requests.html", context)


def _can_create_stock_request(user):
    """Check if user can create stock requests (branch managers with branch assignment)."""
    from .context_processors import get_branch_user_info
    is_branch_user, user_branch_ids = get_branch_user_info(user)
    return is_branch_user and len(user_branch_ids) > 0


@login_required
def create_stock_request(request):
    """
    Create a stock request (Request) for branch managers.
    Branch managers select their branch and items from warehouse inventory to request.
    """
    from django.utils import timezone
    from datetime import datetime
    from .context_processors import get_branch_user_info

    is_branch_user, user_branch_ids = get_branch_user_info(request.user)
    if not is_branch_user:
        messages.error(request, 'You do not have permission to create stock requests. Branch managers only.')
        return redirect('requests')
    if not user_branch_ids:
        messages.error(request, 'You must be assigned to a branch to create stock requests.')
        return redirect('requests')

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            branch_id = data.get('branch_id')
            items = data.get('items', [])  # List of {item_id, quantity}
            notes = (data.get('notes') or '').strip()

            if not branch_id or not items:
                return JsonResponse({'success': False, 'error': 'Select a branch and add at least one item.'}, status=400)

            branch_id = int(branch_id)
            if branch_id not in user_branch_ids:
                return JsonResponse({'success': False, 'error': 'You can only create requests for your assigned branch.'}, status=403)

            branch = get_object_or_404(Branch, id=branch_id, is_active=True)

            # Filter to valid items with quantity > 0
            valid_items = []
            for item_data in items:
                item_id = item_data.get('item_id')
                qty = Decimal(str(item_data.get('quantity', 0)))
                if item_id and qty > 0:
                    valid_items.append((int(item_id), qty))

            if not valid_items:
                return JsonResponse({'success': False, 'error': 'Add at least one item with quantity greater than 0.'}, status=400)

            # Generate request code: REQ-YYYY-NNNNNN (6 digits to differentiate from ItemRequest)
            current_year = datetime.now().year
            last_req = Request.objects.filter(
                request_code__startswith=f'REQ-{current_year}-'
            ).order_by('-request_code').first()

            if last_req:
                try:
                    last_num = int(last_req.request_code.split('-')[-1])
                except (ValueError, IndexError):
                    last_num = 0
                new_num = last_num + 1
            else:
                new_num = 1

            request_code = f'REQ-{current_year}-{new_num:06d}'

            from django.db import transaction
            with transaction.atomic():
                req = Request.objects.create(
                    request_code=request_code,
                    branch=branch,
                    requested_by=request.user,
                    status=Request.StatusType.PENDING,
                    date_of_order=timezone.now(),
                    notes=notes or None,
                )

                for item_id, qty in valid_items:
                    item = get_object_or_404(Item, id=item_id, is_active=True)
                    RequestItem.objects.create(
                        request=req,
                        item=item,
                        qty_requested=qty,
                    )

            messages.success(request, f'Stock request {request_code} created successfully.')
            return JsonResponse({
                'success': True,
                'request_code': request_code,
                'redirect_url': '/requests/',
            })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    # GET - show form (branch managers must have branch assignments)
    user_branches = Branch.objects.filter(
        id__in=user_branch_ids,
        is_active=True
    ).select_related('brand').order_by('brand__name', 'name')

    brands_branches = []
    for branch in user_branches:
        brand_name = branch.brand.name
        existing = next((g for g in brands_branches if g['name'] == brand_name), None)
        if existing:
            existing['branches'].append({'id': branch.id, 'name': branch.name})
        else:
            brands_branches.append({
                'name': brand_name,
                'branches': [{'id': branch.id, 'name': branch.name}],
            })

    # Items: active items with warehouse stock info, filtered to those used in user's branches
    from django.db.models import Sum
    warehouse_stock_qs = StockBalance.objects.filter(
        location__type='WAREHOUSE',
        item__is_active=True
    ).values('item_id').annotate(total=Sum('qty_on_hand'))
    warehouse_stock = {row['item_id']: row['total'] or Decimal('0') for row in warehouse_stock_qs}

    items_list = []
    for item in Item.objects.filter(
        is_active=True, branches__id__in=user_branch_ids
    ).distinct().select_related('brand').prefetch_related('branches').order_by('item_code'):
        qty_available = warehouse_stock.get(item.id, Decimal('0'))
        branch_ids = list(item.branches.filter(is_active=True).values_list('id', flat=True))
        items_list.append({
            'id': item.id,
            'item_code': item.item_code,
            'name': item.name,
            'base_unit': str(item.base_unit) if item.base_unit else '',
            'min_order_qty': float(item.min_stock_qty) if item.min_stock_qty else 0,
            'qty_available': float(qty_available),
            'branch_ids': branch_ids,
        })

    context = {
        'brands_branches': brands_branches,
        'items_list': items_list,
    }
    return render(request, 'maainventory/create_stock_request.html', context)


@login_required
def view_request(request, request_id):
    """View stock request details (read-only). Branch users can only view requests for their branch(es)."""
    from django.http import HttpResponseForbidden
    from .context_processors import get_branch_user_info

    req = get_object_or_404(
        Request.objects.select_related('branch', 'branch__brand', 'requested_by', 'approved_by'),
        id=request_id
    )

    # Branch managers can only view requests for their assigned branch(es)
    is_branch_user, user_branch_ids = get_branch_user_info(request.user)
    if is_branch_user and (not user_branch_ids or req.branch_id not in user_branch_ids):
        return HttpResponseForbidden('You do not have access to this request.')
    request_items = RequestItem.objects.filter(
        request=req
    ).select_related('item', 'variation').prefetch_related('item__photos').order_by('id')

    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_warehouse_staff = user_role and 'Warehouse' in user_role
    is_logistics_staff = user_role and 'Logistics' in user_role

    # Warehouse can mark "Ready for Delivery" (deduct stock, set Ready for Delivery) when status is Warehouse Processing
    can_start_fulfillment = req.status == 'WarehouseProcessing' and is_warehouse_staff

    # Logistics can mark "Out for Delivery" when status is Ready for Delivery
    can_mark_out_for_delivery = req.status == 'ReadyForDelivery' and is_logistics_staff

    # Branch manager for this request's branch can mark as Delivered when status is In Process or Out for Delivery
    is_branch_manager_for_this_branch = is_branch_user and user_branch_ids and req.branch_id in user_branch_ids
    can_mark_delivered = req.status in ('InProcess', 'OutForDelivery') and is_branch_manager_for_this_branch

    items_data = []
    for ri in request_items:
        qty_to_fulfill = ri.qty_approved if ri.qty_approved is not None and ri.qty_approved > 0 else ri.qty_requested
        item = ri.item
        image_url = item.photo_url
        if not image_url and item.photos.exists():
            first_photo = item.photos.first()
            if first_photo and first_photo.photo:
                image_url = request.build_absolute_uri(first_photo.photo.url)
        items_data.append({
            'request_item': ri,
            'item': item,
            'variation': ri.variation,
            'qty_requested': ri.qty_requested,
            'qty_approved': ri.qty_approved,
            'qty_fulfilled': ri.qty_fulfilled,
            'qty_to_fulfill': qty_to_fulfill,
            'image_url': image_url,
        })

    approved_by_name = None
    if req.approved_by:
        approved_by_name = (getattr(req.approved_by.profile, 'full_name', None) or req.approved_by.get_full_name() or req.approved_by.username)

    is_procurement = user_role and 'Procurement' in user_role
    can_approve = is_procurement and req.status == 'Pending'

    context = {
        'req': req,
        'items': items_data,
        'requestor': req.requested_by.profile.full_name if hasattr(req.requested_by, 'profile') and req.requested_by.profile.full_name else (req.requested_by.get_full_name() or req.requested_by.username),
        'approved_by_name': approved_by_name,
        'is_warehouse_staff': is_warehouse_staff,
        'can_start_fulfillment': can_start_fulfillment,
        'can_mark_out_for_delivery': can_mark_out_for_delivery,
        'can_mark_delivered': can_mark_delivered,
        'can_approve': can_approve,
    }
    return render(request, 'maainventory/view_request.html', context)


@login_required
def approve_reject_request(request, request_id):
    """
    Approve or reject a stock request. Procurement managers only.
    POST with action=approve or action=reject. For reject, rejected_reason is required.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    if not (user_role and 'Procurement' in user_role):
        return JsonResponse({'success': False, 'error': 'Only procurement managers can approve or reject requests'}, status=403)

    req = get_object_or_404(Request.objects.prefetch_related('items'), id=request_id)

    if req.status != 'Pending':
        return JsonResponse({
            'success': False,
            'error': f'Request must be Pending Procurement Manager Approval to approve/reject. Current status: {req.get_status_display()}'
        }, status=400)

    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        data = {}

    action = (data.get('action') or '').strip().lower()
    if action not in ('approve', 'reject'):
        return JsonResponse({'success': False, 'error': 'Invalid action. Use approve or reject.'}, status=400)

    from django.utils import timezone
    from django.db import transaction

    with transaction.atomic():
        old_status = req.status
        if action == 'approve':
            req.status = Request.StatusType.WAREHOUSE_PROCESSING
            req.approved_by = request.user
            req.approved_at = timezone.now()
            req.rejected_reason = None
            # Set qty_approved = qty_requested for each item if not already set
            for ri in req.items.all():
                if ri.qty_approved is None or ri.qty_approved <= 0:
                    ri.qty_approved = ri.qty_requested
                    ri.save()
        else:
            rejected_reason = (data.get('rejected_reason') or '').strip()
            if not rejected_reason:
                return JsonResponse({'success': False, 'error': 'Rejection reason is required.'}, status=400)
            req.status = Request.StatusType.REJECTED
            req.approved_by = None
            req.approved_at = None
            req.rejected_reason = rejected_reason
        req.save()

        RequestStatusHistory.objects.create(
            request=req,
            old_status=old_status,
            new_status=req.status,
            changed_by=request.user,
            notes=f'Rejection reason: {req.rejected_reason}' if action == 'reject' and req.rejected_reason else None
        )

    if action == 'approve':
        messages.success(request, f'Request {req.request_code} approved.')
    else:
        messages.success(request, f'Request {req.request_code} rejected.')

    return JsonResponse({
        'success': True,
        'message': f'Request {req.request_code} {action}d.',
        'redirect': request.build_absolute_uri(f'/requests/{req.id}/')
    })


@login_required
def mark_request_in_process(request, request_id):
    """
    Warehouse: Mark Ready for Delivery — deduct from warehouse stock, set qty_fulfilled, set status to Ready for Delivery.
    Only warehouse staff. Request must be Warehouse Processing.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    if not (user_role and 'Warehouse' in user_role):
        return JsonResponse({'success': False, 'error': 'Only warehouse staff can mark request as Ready for Delivery'}, status=403)

    from django.db import transaction
    from django.utils import timezone
    from .models import StockBalance, StockLedger, InventoryLocation

    req = get_object_or_404(Request.objects.prefetch_related('items__item', 'items__variation'), id=request_id)

    if req.status != 'WarehouseProcessing':
        return JsonResponse({
            'success': False,
            'error': f'Request must be Warehouse Processing to mark Ready for Delivery. Current status: {req.get_status_display()}'
        }, status=400)

    warehouse_location, _ = InventoryLocation.objects.get_or_create(
        type='WAREHOUSE',
        defaults={'name': 'Main Warehouse'}
    )

    request_items = list(RequestItem.objects.filter(request=req).select_related('item', 'variation'))

    for ri in request_items:
        qty_to_fulfill = ri.qty_approved if ri.qty_approved is not None and ri.qty_approved > 0 else ri.qty_requested
        if qty_to_fulfill <= 0:
            continue
        try:
            balance = StockBalance.objects.get(
                item=ri.item,
                variation=ri.variation,
                location=warehouse_location
            )
            if balance.qty_on_hand < qty_to_fulfill:
                return JsonResponse({
                    'success': False,
                    'error': f'Insufficient warehouse stock for {ri.item.name}. Available: {balance.qty_on_hand}, Required: {qty_to_fulfill}'
                }, status=400)
        except StockBalance.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'No warehouse stock for {ri.item.name}. Required: {qty_to_fulfill}'
            }, status=400)

    try:
        with transaction.atomic():
            for ri in request_items:
                qty_to_fulfill = ri.qty_approved if ri.qty_approved is not None and ri.qty_approved > 0 else ri.qty_requested
                if qty_to_fulfill <= 0:
                    continue

                balance = StockBalance.objects.get(
                    item=ri.item,
                    variation=ri.variation,
                    location=warehouse_location
                )
                balance.qty_on_hand -= qty_to_fulfill
                balance.save()

                StockLedger.objects.create(
                    item=ri.item,
                    variation=ri.variation,
                    from_location=warehouse_location,
                    qty_change=-qty_to_fulfill,
                    reason='REQUEST_FULFILLMENT',
                    reference_type='REQUEST',
                    reference_id=str(req.id),
                    notes=f'Fulfilled request {req.request_code} to {req.branch.name}',
                    created_by=request.user
                )

                ri.qty_fulfilled = qty_to_fulfill
                ri.save()

            old_status = req.status
            req.status = Request.StatusType.READY_FOR_DELIVERY
            req.save()

            RequestStatusHistory.objects.create(
                request=req,
                old_status=old_status,
                new_status=req.status,
                changed_by=request.user,
                notes='Warehouse marked Ready for Delivery; inventory deducted from warehouse.'
            )

        messages.success(request, f'Request {req.request_code} is now Ready for Delivery. Inventory deducted from warehouse.')
        return JsonResponse({
            'success': True,
            'message': f'Request {req.request_code} is now Ready for Delivery.'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def mark_request_out_for_delivery(request, request_id):
    """
    Logistics: Mark request as Out for Delivery. Only logistics staff. Request must be Ready for Delivery.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    if not (user_role and 'Logistics' in user_role):
        return JsonResponse({'success': False, 'error': 'Only logistics staff can mark requests as Out for Delivery'}, status=403)

    req = get_object_or_404(Request.objects.prefetch_related('items'), id=request_id)

    if req.status != 'ReadyForDelivery':
        return JsonResponse({
            'success': False,
            'error': f'Request must be Ready for Delivery to mark as Out for Delivery. Current status: {req.get_status_display()}'
        }, status=400)

    from django.db import transaction

    try:
        with transaction.atomic():
            old_status = req.status
            req.status = Request.StatusType.OUT_FOR_DELIVERY
            req.save()

            RequestStatusHistory.objects.create(
                request=req,
                old_status=old_status,
                new_status=req.status,
                changed_by=request.user,
                notes='Logistics marked Out for Delivery.'
            )

        messages.success(request, f'Request {req.request_code} is now Out for Delivery.')
        return JsonResponse({
            'success': True,
            'message': f'Request {req.request_code} is now Out for Delivery. Branch manager can mark as Delivered when items arrive.'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def mark_request_delivered(request, request_id):
    """
    Mark request as delivered. Branch Manager only (for the branch that requested the items).
    Does not deduct inventory — warehouse already did that when marking In Process.
    When delivered, request items appear on the Branches page for that branch.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    from .context_processors import get_branch_user_info

    is_branch_user, user_branch_ids = get_branch_user_info(request.user)
    if not is_branch_user or not user_branch_ids:
        return JsonResponse({'success': False, 'error': 'Only the branch manager for this branch can mark the request as delivered'}, status=403)

    req = get_object_or_404(Request.objects.prefetch_related('items__item', 'items__variation'), id=request_id)

    if req.branch_id not in user_branch_ids:
        return JsonResponse({'success': False, 'error': 'You can only mark as delivered for requests belonging to your branch'}, status=403)

    if req.status in ('Delivered', 'Completed'):
        return JsonResponse({'success': False, 'error': 'Request has already been delivered'})

    if req.status not in ('InProcess', 'OutForDelivery'):
        return JsonResponse({
            'success': False,
            'error': f'Request must be In Process or Out for Delivery to mark as delivered. Current status: {req.get_status_display()}'
        }, status=400)

    from django.db import transaction
    from django.utils import timezone

    try:
        with transaction.atomic():
            old_status = req.status
            req.status = Request.StatusType.DELIVERED
            req.save()

            RequestStatusHistory.objects.create(
                request=req,
                old_status=old_status,
                new_status=req.status,
                changed_by=request.user,
                notes='Branch manager confirmed delivery. Items will appear on the Branches page for this branch.'
            )

        messages.success(request, f'Request {req.request_code} marked as delivered. Items are now shown for branch "{req.branch.name}" on the Branches page.')
        return JsonResponse({
            'success': True,
            'message': f'Request {req.request_code} marked as delivered. Items appear on the Branches page for {req.branch.name}.',
            'branches_url': request.build_absolute_uri('/branches/')
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def new_request(request):
    """Render 'New Stock Order' page with branches, suppliers, and items."""
    from .models import Branch, Supplier, SupplierItem, SupplierStock, SupplierOrder, SupplierOrderItem, PortalToken
    import json
    from django.utils import timezone
    from datetime import datetime
    from decimal import Decimal
    from django.db.models import Sum
    
    # Handle POST request (Place Order)
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            branch_id = data.get('branch_id')
            supplier_id = data.get('supplier_id')
            items = data.get('items', [])  # List of {code, name, quantity, price}
            
            if not branch_id or not supplier_id or not items:
                return JsonResponse({'success': False, 'error': 'Missing required data'}, status=400)
            
            # Get branch and supplier
            branch = get_object_or_404(Branch, id=branch_id, is_active=True)
            supplier = get_object_or_404(Supplier, id=supplier_id, is_active=True)
            
            # Generate PO code (PO-YYYY######)
            current_year = datetime.now().year
            last_order = SupplierOrder.objects.filter(
                po_code__startswith=f'PO-{current_year}'
            ).order_by('-po_code').first()
            
            if last_order:
                # Extract number from PO code
                # Handle both formats: PO-2026-0001 (old) and PO-20260001 (new)
                po_code_str = last_order.po_code
                if '-' in po_code_str and po_code_str.count('-') == 2:
                    # Old format: PO-2026-0001
                    last_num = int(po_code_str.split('-')[-1])
                else:
                    # New format: PO-20260001 (last 6 digits)
                    last_6_digits = po_code_str[-6:] if len(po_code_str) >= 6 else po_code_str
                    last_num = int(last_6_digits)
                new_num = last_num + 1
            else:
                new_num = 1
            
            po_code = f'PO-{current_year}{new_num:06d}'
            
            # Create Supplier Order
            supplier_order = SupplierOrder.objects.create(
                po_code=po_code,
                supplier=supplier,
                created_by=request.user,
                status='Sent'  # Set to SENT since email is sent immediately
            )
            
            # Generate secure token for supplier access
            import secrets
            from django.utils import timezone
            from django.conf import settings
            from datetime import timedelta
            
            token_string = secrets.token_urlsafe(32)  # 32-character URL-safe token
            # No expiration - tokens are valid indefinitely
            expires_at = timezone.now() + timedelta(days=3650)  # 10 years (effectively no expiration)
            
            portal_token = PortalToken.objects.create(
                token=token_string,
                supplier=supplier,
                supplier_order=supplier_order,
                expires_at=expires_at
            )
            
            # First, validate all items have sufficient stock
            for item_data in items:
                item_code = item_data.get('code')
                quantity = Decimal(str(item_data.get('quantity', 0)))
                
                if quantity <= 0:
                    continue
                
                # Get the item
                item = Item.objects.filter(item_code=item_code, is_active=True).first()
                if not item:
                    return JsonResponse({
                        'success': False, 
                        'error': f'Item {item_code} not found'
                    }, status=400)
                
                # Check available stock
                available_stock = SupplierStock.objects.filter(
                    supplier=supplier,
                    item=item
                ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
                
                if available_stock < quantity:
                    return JsonResponse({
                        'success': False, 
                        'error': f'Insufficient stock for {item.name}. Available: {available_stock}, Requested: {quantity}'
                    }, status=400)
            
            # Create Order Items and deduct from stock
            for item_data in items:
                item_code = item_data.get('code')
                quantity = Decimal(str(item_data.get('quantity', 0)))
                price = item_data.get('price', 0)
                
                if quantity <= 0:
                    continue
                
                # Get the item
                item = Item.objects.filter(item_code=item_code, is_active=True).first()
                if not item:
                    continue
                
                # Create order item
                SupplierOrderItem.objects.create(
                    supplier_order=supplier_order,
                    item=item,
                    qty_ordered=quantity,
                    price_per_unit=price
                )
                
                # Deduct from supplier stock
                remaining_qty = quantity
                stock_entries = SupplierStock.objects.filter(
                    supplier=supplier,
                    item=item,
                    quantity__gt=0
                ).order_by('confirmed_at')  # Use oldest stock first (FIFO)
                
                for stock_entry in stock_entries:
                    if remaining_qty <= 0:
                        break
                    
                    if stock_entry.quantity >= remaining_qty:
                        # This entry has enough stock
                        stock_entry.quantity -= remaining_qty
                        remaining_qty = Decimal('0')
                        if stock_entry.quantity == 0:
                            stock_entry.delete()  # Remove entry if quantity becomes 0
                        else:
                            stock_entry.save()
                    else:
                        # Use all of this entry and continue to next
                        remaining_qty -= stock_entry.quantity
                        stock_entry.delete()  # Remove depleted entry
            
            # Send email to supplier with invoice
            try:
                send_invoice_email(supplier_order, request)
            except Exception as email_error:
                # Log error but don't fail the order creation
                print(f"Error sending email: {email_error}")
            
            return JsonResponse({
                'success': True,
                'po_code': po_code,
                'order_id': supplier_order.id
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    # GET request - show the form
    # Group branches by brand, with Warehouse first
    brands_branches = []
    all_brands = list(Brand.objects.all())
    
    # Sort brands: Warehouse first, then alphabetically
    def brand_sort_key(brand):
        if brand.name.lower() == 'warehouse':
            return (0, brand.name.lower())
        return (1, brand.name.lower())
    
    sorted_brands = sorted(all_brands, key=brand_sort_key)
    
    for brand in sorted_brands:
        branches = brand.branches.filter(is_active=True).order_by('name')
        if branches.exists():
            brands_branches.append({
                "id": brand.id,
                "name": brand.name,
                "branches": [{"id": b.id, "name": b.name} for b in branches]
            })
    
    # Get all active suppliers with their items
    suppliers = Supplier.objects.filter(is_active=True).select_related('category').prefetch_related(
        'supplier_items__item__branches'
    ).order_by('name')
    
    # Build suppliers list with branch information
    suppliers_list = []
    for supplier in suppliers:
        # Get all items from this supplier and their associated branches
        supplier_items = SupplierItem.objects.filter(
            supplier=supplier,
            is_active=True
        ).select_related('item').prefetch_related('item__branches')
        
        # Collect all unique branches that use items from this supplier
        branch_ids = set()
        for supplier_item in supplier_items:
            if supplier_item.item and supplier_item.item.branches.exists():
                branch_ids.update(supplier_item.item.branches.values_list('id', flat=True))
        
        suppliers_list.append({
            "id": supplier.id,
            "name": supplier.name,
            "category": supplier.category.name if supplier.category else "",
            "branches": list(branch_ids),  # List of branch IDs that use this supplier's items
        })
    
    # Calculate pending quantities from purchase orders
    from .models import SupplierOrder, SupplierOrderItem
    pending_orders = SupplierOrder.objects.exclude(
        status__in=['Received', 'Cancelled']
    ).select_related('supplier')
    
    # Create a dictionary to store pending quantities by (supplier_id, item_id)
    pending_quantities = {}
    for order in pending_orders:
        order_items = SupplierOrderItem.objects.filter(
            supplier_order=order
        ).select_related('item')
        
        for order_item in order_items:
            key = (order.supplier.id, order_item.item.id)
            pending_qty = order_item.qty_ordered - order_item.qty_received
            if key not in pending_quantities:
                pending_quantities[key] = Decimal('0.00')
            pending_quantities[key] += pending_qty
    
    # Get all items with their suppliers and branches - ONLY from Supplier Stock
    items_list = []
    
    # Get items that are currently in supplier stock
    from .models import SupplierStock
    from django.db.models import Q
    
    stock_items = SupplierStock.objects.filter(
        quantity__gt=0,
        supplier__is_active=True,
        item__is_active=True
    ).select_related('supplier', 'item', 'item__brand').prefetch_related('item__branches')
    
    for stock_item in stock_items:
        item = stock_item.item
        supplier = stock_item.supplier
        
        # Get branches that use this item
        item_branches = item.branches.filter(is_active=True).values_list('id', flat=True)
        
        # Get pending quantity for this supplier/item combination
        pending_key = (supplier.id, item.id)
        pending_qty = float(pending_quantities.get(pending_key, Decimal('0.00')))
        
        # Get price per unit from SupplierItem or Item
        price_per_unit = None
        supplier_item = SupplierItem.objects.filter(
            item=item,
            supplier=supplier,
            is_active=True
        ).first()
        
        if supplier_item and supplier_item.price_per_unit:
            price_per_unit = float(supplier_item.price_per_unit)
        elif item.price_per_unit:
            price_per_unit = float(item.price_per_unit)
        
        items_list.append({
            "code": item.item_code,
            "name": item.name,
            "supplier_id": supplier.id,
            "supplier_name": supplier.name,
            "branches": list(item_branches),  # List of branch IDs that use this item
            "price_per_unit": price_per_unit,
            "available_quantity": float(stock_item.quantity),  # Show available quantity
            "pending_quantity": pending_qty,  # Show pending quantity
        })

    context = {
        "brands_branches": brands_branches,
        "suppliers": suppliers_list,
        "items": items_list,
    }
    return render(request, "maainventory/new_request.html", context)


@login_required
def add_item(request):
    """Render 'Add Supplier Item' page (similar to new_request page structure)."""
    # Get all active suppliers with their categories
    suppliers = Supplier.objects.filter(is_active=True).select_related('category').order_by('name')
    
    # Get supplier categories
    categories = SupplierCategory.objects.filter(is_active=True).order_by('name')
    
    # Build brands and branches structure
    brands_branches = []
    for brand in Brand.objects.all().order_by('name'):
        branches = brand.branches.filter(is_active=True).order_by('name')
        if branches.exists():
            brands_branches.append({
                "id": brand.id,
                "name": brand.name,
                "branches": [{"id": b.id, "name": b.name} for b in branches]
            })
    
    # Handle form submission
    if request.method == 'POST':
        # Get supplier from hidden input if dropdown is disabled
        supplier_id = request.POST.get('supplier') or request.POST.get('supplier-hidden')
        if supplier_id:
            # Create a mutable copy of POST data
            post_data = request.POST.copy()
            post_data['supplier'] = supplier_id
            form = SupplierItemForm(post_data)
        else:
            form = SupplierItemForm(request.POST)
        
        if form.is_valid():
            photos = request.FILES.getlist('photos')
            if not photos:
                form.add_error(None, 'At least one photo is required.')
            else:
                supplier_item = form.save(commit=True, user=request.user)
                # Save many-to-many branches to Item (not SupplierItem)
                branch_ids = request.POST.getlist('branches')
                if branch_ids and supplier_item.item:
                    supplier_item.item.branches.set(branch_ids)
                
                # Handle photo uploads (maximum 5)
                if photos and supplier_item and supplier_item.item:
                    existing_count = ItemPhoto.objects.filter(item=supplier_item.item).count()
                    remaining_slots = 5 - existing_count
                    if remaining_slots > 0:
                        photos_to_save = photos[:remaining_slots]
                        saved_count = 0
                        for idx, photo in enumerate(photos_to_save):
                            try:
                                ItemPhoto.objects.create(
                                    item=supplier_item.item,
                                    photo=photo,
                                    order=existing_count + idx
                                )
                                saved_count += 1
                            except Exception as e:
                                messages.error(request, f'Error saving photo {photo.name}: {str(e)}')
                        
                        if saved_count > 0:
                            messages.success(request, f'{saved_count} photo(s) saved successfully.')
                        
                        if len(photos) > remaining_slots:
                            messages.warning(request, f'Only {remaining_slots} photo(s) were saved. Maximum 5 photos per item.')
                    else:
                        messages.warning(request, 'Maximum 5 photos already exist for this item.')
                
                messages.success(request, f'Supplier item "{supplier_item.item.name}" for "{supplier_item.supplier.name}" added successfully.')
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.urls import reverse
                    return JsonResponse({'success': True, 'redirect_url': reverse('inventory')})
                return redirect('inventory')
        if not form.is_valid():
            error_parts = []
            for err in form.non_field_errors():
                error_parts.append(err)
            for name in form.errors:
                if name in (None, '__all__') or name not in form.fields:
                    continue
                label = form.fields[name].label
                errs = form.errors[name]
                first_err = errs[0] if errs else 'Invalid'
                error_parts.append(f"{label}: {first_err}")
            if error_parts:
                msg = "Please correct the errors below:\n" + "\n".join("- " + part for part in error_parts)
            else:
                msg = "Please correct the errors below."
            messages.error(request, msg)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors_dict = {}
                for name in form.errors:
                    if name in (None, '__all__'):
                        errors_dict['__all__'] = list(form.errors[name])
                    elif name in form.fields:
                        errors_dict[name] = list(form.errors[name])
                return JsonResponse({'success': False, 'message': msg, 'errors': errors_dict})
    else:
        form = SupplierItemForm()
    
    # Build suppliers list for left sidebar with category information
    suppliers_list = []
    for supplier in suppliers:
        suppliers_list.append({
            "id": supplier.id,
            "name": supplier.name,
            "category": supplier.category.name if supplier.category else None,
            "category_id": supplier.category.id if supplier.category else None,
        })
    
    # When re-rendering after validation error, pass selected supplier so UI can restore it
    selected_supplier_id = None
    selected_supplier_category_id = None
    if request.method == 'POST':
        sid = request.POST.get('supplier') or request.POST.get('supplier-hidden')
        if sid:
            try:
                selected_supplier_id = int(sid)
                for s in suppliers_list:
                    if s['id'] == selected_supplier_id:
                        selected_supplier_category_id = s.get('category_id')
                        break
            except (ValueError, TypeError):
                pass
    
    # Get all active items for reference (not used in form, but available if needed)
    items = Item.objects.filter(is_active=True).select_related('brand').order_by('item_code')
    
    context = {
        "form": form,
        "suppliers": suppliers_list,
        "categories": categories,
        "brands_branches": brands_branches,
        "items": items,  # Available for reference/autocomplete if needed
        "selected_supplier_id": selected_supplier_id,
        "selected_supplier_category_id": selected_supplier_category_id,
    }
    return render(request, "maainventory/add_item.html", context)


@login_required
def edit_supplier(request, code):
    """Render edit page for a single supplier with dynamic data and handle form submission."""
    # Parse supplier code (format: SUP-001) to extract ID
    try:
        if code.startswith('SUP-'):
            supplier_id = int(code.replace('SUP-', ''))
        else:
            supplier_id = int(code)
        supplier = get_object_or_404(Supplier, id=supplier_id)
    except (ValueError, AttributeError):
        raise Http404("Supplier not found")

    # Check if user has permission (Procurement Manager or IT)
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_procurement = user_role and 'Procurement' in user_role
    is_it = user_role and ('IT' in user_role or user_role == 'IT')
    
    if not (is_procurement or is_it):
        messages.error(request, 'You do not have permission to edit suppliers.')
        return redirect('suppliers')

    # Handle form submission
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, f'Supplier "{supplier.name}" updated successfully.')
            return redirect('edit_supplier', code=code)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SupplierForm(instance=supplier)

    # Get supplier items
    supplier_items = SupplierItem.objects.filter(
        supplier=supplier,
        is_active=True
    ).select_related('item', 'variation')
    
    # Format supplier items for template
    supplier_items_list = []
    for item in supplier_items:
        # Get last order date for this item
        last_order = SupplierOrder.objects.filter(
            supplier=supplier,
            items__item=item.item,
            items__variation=item.variation
        ).order_by('-created_at').first()
        
        supplier_items_list.append({
            "item_code": item.item.item_code,
            "item_name": item.item.name,
            "variation": item.variation.variation_name if item.variation else None,
            "price_per_unit": f"{item.price_per_unit:.2f}",
            "last_purchase_date": last_order.created_at.strftime("%Y-%m-%d") if last_order else "—",
            "status": "Active",
        })
    
    # Get supplier orders
    orders = SupplierOrder.objects.filter(supplier=supplier).order_by('-created_at')[:10]
    
    # Format orders for template
    orders_list = []
    for order in orders:
        orders_list.append({
            "po_code": order.po_code,
            "status": order.get_status_display(),
            "created_at": order.created_at.strftime("%Y-%m-%d"),
            "requested_delivery_date": order.requested_delivery_date.strftime("%Y-%m-%d") if order.requested_delivery_date else "—",
        })
    
    # Get items held at supplier
    supplier_hold_location = InventoryLocation.objects.filter(
        type='SUPPLIER_HOLD',
        supplier=supplier
    ).first()
    
    supplier_hold_items = []
    if supplier_hold_location:
        hold_stocks = StockBalance.objects.filter(
            location=supplier_hold_location
        ).select_related('item', 'variation')
        for stock in hold_stocks:
            if stock.qty_on_hand > 0:
                supplier_hold_items.append({
                    "item_code": stock.item.item_code,
                    "item_name": stock.item.name,
                    "variation": stock.variation.variation_name if stock.variation else None,
                    "qty": float(stock.qty_on_hand),
                    "unit": stock.item.base_unit,
                })
    
    # Build supplier data for template
    supplier_data = {
        "code": f"SUP-{supplier.id:03d}",
        "name": supplier.name,
        "email": supplier.email,
        "phone": supplier.phone,
        "address": supplier.address or "",
        "category": supplier.category.name if supplier.category else "—",
        "category_id": supplier.category.id if supplier.category else None,
        "contact_person": supplier.contact_person or "",
        "delivery_days": supplier.delivery_days if isinstance(supplier.delivery_days, dict) else {},
        "supplier_items": supplier_items_list,
        "orders": orders_list,
        "supplier_hold_items": supplier_hold_items,
        "is_active": supplier.is_active,
    }

    # Get all active categories for dropdown
    categories = SupplierCategory.objects.filter(is_active=True).order_by('name')

    context = {
        "item": supplier_data,
        "form": form,
        "supplier": supplier,
        "categories": categories,
    }
    return render(request, "maainventory/edit_supplier.html", context)


@login_required
def add_supplier(request):
    """Add a new supplier (Procurement Manager only)"""
    # Check if user is Procurement Manager
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_procurement = user_role and 'Procurement' in user_role
    
    if not is_procurement:
        messages.error(request, 'You do not have permission to perform this action.')
        return redirect('suppliers')
    
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.created_by = request.user
            supplier.save()
            messages.success(request, f'Supplier "{supplier.name}" added successfully.')
            return redirect('suppliers')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SupplierForm()
    
    context = {
        'form': form,
        'title': 'Add New Supplier',
    }
    return render(request, 'maainventory/add_supplier.html', context)


@login_required
def suppliers(request):
    """Render suppliers page with dynamic data from database."""
    from django.core.paginator import Paginator
    
    # Check if user is Procurement Manager to show add button
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    
    # Block warehouse staff from accessing suppliers page
    if user_role and 'Warehouse' in user_role:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('dashboard')
    
    is_procurement = user_role and 'Procurement' in user_role
    
    # Get all active suppliers
    suppliers_queryset = Supplier.objects.filter(is_active=True).annotate(
        total_items=Count('supplier_items', filter=Q(supplier_items__is_active=True)),
        active_orders_count=Count('orders', filter=Q(orders__status__in=['Draft', 'Sent', 'Confirmed', 'InProduction', 'Ready', 'PartiallyReceived'])),
    ).order_by('name')
    
    # Build suppliers list for template
    suppliers_list = []
    for supplier in suppliers_queryset:
        # Get total spend from completed orders (simplified - would need SupplierSpendMonthly or calculate from orders)
        # For now, we'll use a placeholder
        total_spend = "0.00"  # TODO: Calculate from SupplierOrder history
        
        # Get pending invoices count (orders with invoice signatures but not fully received)
        pending_invoices = SupplierOrder.objects.filter(
            supplier=supplier,
            invoice_signatures__isnull=False
        ).exclude(status='Received').count()
        
        # Get items held at supplier
        supplier_hold_location = InventoryLocation.objects.filter(
            type='SUPPLIER_HOLD',
            supplier=supplier
        ).first()
        
        supplier_hold_qty = "0"
        if supplier_hold_location:
            hold_total = StockBalance.objects.filter(
                location=supplier_hold_location
            ).aggregate(
                total=Coalesce(Sum('qty_on_hand'), Decimal('0'), output_field=DecimalField())
            )['total'] or Decimal('0')
            supplier_hold_qty = f"{hold_total:,.0f}" if hold_total > 0 else "No"
        
        # Get last order date
        last_order = SupplierOrder.objects.filter(
            supplier=supplier
        ).order_by('-created_at').first()
        
        last_ordered_date = last_order.created_at.date() if last_order else None
        
        suppliers_list.append({
            "code": f"SUP-{supplier.id:03d}",
            "name": supplier.name,
            "email": supplier.email,
            "phone": supplier.phone,
            "address": supplier.address or "",
            "category": supplier.category.name if supplier.category else "",
            "category_id": supplier.category.id if supplier.category else None,
            "contact_person": supplier.contact_person or "—",
            "delivery_days": supplier.delivery_days if isinstance(supplier.delivery_days, str) else (', '.join(supplier.delivery_days.keys()) if isinstance(supplier.delivery_days, dict) else str(supplier.delivery_days)),
            "total_items": supplier.total_items,
            "active_orders": supplier.active_orders_count,
            "total_spend": total_spend,
            "pending_invoices": pending_invoices,
            "supplier_hold": supplier_hold_qty,
            "last_ordered_date": last_ordered_date.strftime("%Y-%m-%d") if last_ordered_date else "—",
            "id": supplier.id,
        })

    # Paginate suppliers
    paginator = Paginator(suppliers_list, 10)  # 10 items per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Get all active categories for dropdown
    categories = SupplierCategory.objects.filter(is_active=True).order_by('name')
    
    # Get supplier items for price discussion modal (grouped by supplier)
    # Get ALL suppliers (not just paginated) for the dropdown
    all_suppliers = Supplier.objects.filter(is_active=True).order_by('name')
    supplier_items_by_supplier = {}
    
    for supplier in all_suppliers:
        items = SupplierItem.objects.filter(
            supplier=supplier,
            is_active=True
        ).select_related('item', 'variation', 'base_unit').order_by('item__item_code')
        
        supplier_items_by_supplier[supplier.id] = [
            {
                'id': item.id,
                'item_code': item.item.item_code,
                'item_name': item.item.name,
                'variation': item.variation.variation_name if item.variation else None,
                'current_price': float(item.price_per_unit),
                'base_unit': item.base_unit.abbreviation if item.base_unit else item.item.base_unit
            }
            for item in items
        ]
    
    # Also create a list of all suppliers for the dropdown
    all_suppliers_list = [
        {'id': s.id, 'name': s.name}
        for s in all_suppliers
    ]
    
    context = {
        "items": page_obj,
        "page_obj": page_obj,
        "is_procurement": is_procurement,
        "categories": categories,
        "supplier_items_json": json.dumps(supplier_items_by_supplier),
        "all_suppliers": all_suppliers_list,
    }

    return render(request, "maainventory/suppliers.html", context)


@login_required
def delete_supplier(request, code):
    """Delete a supplier (Procurement Manager or IT only)"""
    # Parse supplier code (format: SUP-001) to extract ID
    try:
        if code.startswith('SUP-'):
            supplier_id = int(code.replace('SUP-', ''))
        else:
            supplier_id = int(code)
        supplier = get_object_or_404(Supplier, id=supplier_id)
    except (ValueError, AttributeError):
        raise Http404("Supplier not found")
    
    # Check if user has permission (Procurement Manager or IT)
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_procurement = user_role and 'Procurement' in user_role
    is_it = user_role and ('IT' in user_role or user_role == 'IT')
    
    if not (is_procurement or is_it):
        messages.error(request, 'You do not have permission to delete suppliers.')
        return redirect('suppliers')
    
    # Delete the supplier (confirmation already handled in JavaScript)
    supplier_name = supplier.name
    supplier.delete()
    messages.success(request, f'Supplier "{supplier_name}" has been deleted successfully.')
    return redirect('suppliers')


@login_required
def update_supplier_category(request):
    """Update supplier category via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    import json
    
    try:
        data = json.loads(request.body)
        supplier_id = data.get('supplier_id')
        category_id = data.get('category_id')
        
        if not supplier_id:
            return JsonResponse({'success': False, 'error': 'Supplier ID is required'}, status=400)
        
        supplier = get_object_or_404(Supplier, id=supplier_id)
        
        # Check if user has permission (Procurement Manager or IT)
        user_profile = getattr(request.user, 'profile', None)
        user_role = user_profile.role.name if user_profile and user_profile.role else None
        is_procurement = user_role and 'Procurement' in user_role
        is_it = user_role and ('IT' in user_role or user_role == 'IT')
        
        if not (is_procurement or is_it):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        # Update category
        if category_id:
            category = get_object_or_404(SupplierCategory, id=category_id, is_active=True)
            supplier.category = category
        else:
            supplier.category = None
        
        supplier.save()
        
        return JsonResponse({
            'success': True,
            'category_name': supplier.category.name if supplier.category else None
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def register(request):
    """Handle user registration"""
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save(commit=True)  # Explicitly save to database
                # Verify the user and profile were created
                from .models import UserProfile
                profile = UserProfile.objects.get(user=user)
                messages.success(request, f'Account created successfully for {user.username}! Please log in to continue.')
                return redirect('login')
            except Exception as e:
                messages.error(request, f'An error occurred during registration: {str(e)}')
                # Log the error for debugging
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Registration error: {str(e)}')
    else:
        form = RegistrationForm()
    
    # Load all active branches for JavaScript filtering
    from .models import Branch
    all_branches = Branch.objects.filter(is_active=True).select_related('brand').order_by('brand__name', 'name')
    
    # Update form's branches queryset to include all branches for template rendering
    form.fields['branches'].queryset = all_branches
    
    return render(request, 'maainventory/register.html', {
        'form': form,
        'all_branches': all_branches
    })


@login_required
def punch_id_management(request):
    """IT-only page to manage valid Punch IDs"""
    # Check if user is IT
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_it = user_role and ('IT' in user_role or user_role == 'IT')
    
    if not is_it:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('dashboard')
    
    # Get all valid punch IDs
    punch_ids = ValidPunchID.objects.all().order_by('punch_id')
    
    # Check which ones are already used
    used_punch_ids = set(
        UserProfile.objects.exclude(punch_id__isnull=True)
        .exclude(punch_id='')
        .values_list('punch_id', flat=True)
    )
    
    context = {
        'punch_ids': punch_ids,
        'used_punch_ids': used_punch_ids,
    }
    
    return render(request, 'maainventory/punch_id_management.html', context)


@login_required
def punch_id_add(request):
    """Add a new valid Punch ID (IT only)"""
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_it = user_role and ('IT' in user_role or user_role == 'IT')
    
    if not is_it:
        messages.error(request, 'You do not have permission to perform this action.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        punch_id = request.POST.get('punch_id', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        if not punch_id:
            messages.error(request, 'Punch ID is required.')
            return redirect('punch_id_management')
        
        # Check if punch ID already exists in ValidPunchID table
        if ValidPunchID.objects.filter(punch_id=punch_id).exists():
            messages.error(request, f'Punch ID "{punch_id}" already exists in the valid list.')
            return redirect('punch_id_management')
        
        try:
            ValidPunchID.objects.create(
                punch_id=punch_id,
                is_active=is_active,
                created_by=request.user
            )
            messages.success(request, f'Punch ID "{punch_id}" added successfully.')
        except Exception as e:
            messages.error(request, f'Error adding Punch ID: {str(e)}')
    
    return redirect('punch_id_management')


@login_required
def punch_id_edit(request, punch_id_id):
    """Edit an existing valid Punch ID (IT only)"""
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_it = user_role and ('IT' in user_role or user_role == 'IT')
    
    if not is_it:
        messages.error(request, 'You do not have permission to perform this action.')
        return redirect('dashboard')
    
    valid_punch_id = get_object_or_404(ValidPunchID, id=punch_id_id)
    
    if request.method == 'GET':
        # Show edit form
        context = {
            'punch_id': valid_punch_id,
        }
        return render(request, 'maainventory/punch_id_edit.html', context)
    
    if request.method == 'POST':
        new_punch_id = request.POST.get('punch_id', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        if not new_punch_id:
            messages.error(request, 'Punch ID is required.')
            return redirect('punch_id_management')
        
        # Check if new punch ID conflicts with existing one in ValidPunchID table (excluding current)
        if ValidPunchID.objects.filter(punch_id=new_punch_id).exclude(id=punch_id_id).exists():
            messages.error(request, f'Punch ID "{new_punch_id}" already exists in the valid list.')
            return redirect('punch_id_management')
        
        try:
            valid_punch_id.punch_id = new_punch_id
            valid_punch_id.is_active = is_active
            valid_punch_id.save()
            messages.success(request, f'Punch ID updated successfully.')
        except Exception as e:
            messages.error(request, f'Error updating Punch ID: {str(e)}')
    
    return redirect('punch_id_management')


@login_required
def punch_id_delete(request, punch_id_id):
    """Delete a valid Punch ID (IT only)"""
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_it = user_role and ('IT' in user_role or user_role == 'IT')
    
    if not is_it:
        messages.error(request, 'You do not have permission to perform this action.')
        return redirect('dashboard')
    
    valid_punch_id = get_object_or_404(ValidPunchID, id=punch_id_id)
    
    # Check if this punch ID is already registered
    if UserProfile.objects.filter(punch_id=valid_punch_id.punch_id).exists():
        messages.error(request, f'Cannot delete Punch ID "{valid_punch_id.punch_id}" because it is already registered by a user.')
        return redirect('punch_id_management')
    
    try:
        punch_id_value = valid_punch_id.punch_id
        valid_punch_id.delete()
        messages.success(request, f'Punch ID "{punch_id_value}" deleted successfully.')
    except Exception as e:
        messages.error(request, f'Error deleting Punch ID: {str(e)}')
    
    return redirect('punch_id_management')


@login_required
@login_required
def purchase_orders(request):
    """Render purchase orders page with all submitted supplier orders"""
    from django.core.paginator import Paginator
    
    # Check if user is warehouse staff
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_warehouse_staff = user_role and 'Warehouse' in user_role
    
    # Get all supplier orders, ordered by most recent
    orders_queryset = SupplierOrder.objects.select_related(
        'supplier', 'created_by'
    ).prefetch_related('items__item').order_by('-created_at')
    
    # Build orders list for template
    orders_list = []
    for order in orders_queryset:
        # Calculate total items, quantity, and amount
        total_items = order.items.count()
        total_quantity = sum(item.qty_ordered for item in order.items.all())
        total_amount = sum(item.qty_ordered * item.price_per_unit for item in order.items.all())
        
        # Get item names for display
        item_names = [item.item.name for item in order.items.all()[:3]]  # First 3 items
        if total_items > 3:
            item_names.append(f"+ {total_items - 3} more")
        
        orders_list.append({
            "id": order.id,
            "po_code": order.po_code,
            "supplier": order.supplier.name,
            "created_by": order.created_by.profile.full_name if hasattr(order.created_by, 'profile') and order.created_by.profile.full_name else (order.created_by.get_full_name() or order.created_by.username),
            "created_at": order.created_at.strftime("%Y-%m-%d"),
            "created_at_display": order.created_at.strftime("%B %d, %Y"),
            "status": order.get_status_display(),
            "status_value": order.status,
            "total_items": total_items,
            "total_quantity": total_quantity,
            "total_amount": f"{total_amount:,.2f}",
            "item_names": ", ".join(item_names),
            "requested_delivery_date": order.requested_delivery_date.strftime("%Y-%m-%d") if order.requested_delivery_date else None,
            "email_sent_at": order.email_sent_at.strftime("%Y-%m-%d %H:%M") if order.email_sent_at else None,
        })
    
    # Paginate orders
    paginator = Paginator(orders_list, 10)  # 10 items per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        "orders": page_obj,
        "page_obj": page_obj,
        "is_warehouse_staff": is_warehouse_staff,
    }
    
    return render(request, "maainventory/purchase_orders.html", context)


@login_required
def mark_order_received(request, order_id):
    """Mark purchase order as received and add items to inventory (Warehouse Staff only)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    
    # Check if user is warehouse staff
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    if not (user_role and 'Warehouse' in user_role):
        return JsonResponse({'success': False, 'error': 'Only warehouse staff can mark orders as received'}, status=403)
    
    try:
        import json
        from django.db import transaction
        from .models import SupplierOrder, SupplierOrderItem, StockBalance, StockLedger, InventoryLocation
        from django.utils import timezone
        
        # Get note from request body if provided
        note = ''
        try:
            data = json.loads(request.body)
            note = data.get('note', '').strip()
        except (json.JSONDecodeError, AttributeError):
            pass
        
        order = get_object_or_404(SupplierOrder, id=order_id)
        
        # Check if already received
        if order.status == 'Received':
            return JsonResponse({'success': False, 'error': 'Order has already been marked as received'})
        
        # Get warehouse location
        warehouse_location, created = InventoryLocation.objects.get_or_create(
            type='WAREHOUSE',
            defaults={'name': 'Main Warehouse'}
        )
        
        with transaction.atomic():
            # Get all order items
            order_items = SupplierOrderItem.objects.filter(supplier_order=order).select_related('item', 'variation')
            
            # Process each item
            for order_item in order_items:
                # Update or create stock balance
                stock_balance, created = StockBalance.objects.get_or_create(
                    item=order_item.item,
                    variation=order_item.variation,
                    location=warehouse_location,
                    defaults={'qty_on_hand': Decimal('0.00')}
                )
                
                # Add quantity to stock
                stock_balance.qty_on_hand += order_item.qty_ordered
                stock_balance.save()
                
                # Create ledger entry for audit trail
                ledger_notes = f'Received from PO {order.po_code} - Supplier: {order.supplier.name}'
                if note:
                    ledger_notes += f'\nWarehouse Note: {note}'
                
                StockLedger.objects.create(
                    item=order_item.item,
                    variation=order_item.variation,
                    to_location=warehouse_location,
                    qty_change=order_item.qty_ordered,
                    reason='DELIVERY_RECEIVED',
                    reference_type='SUPPLIER_ORDER',
                    reference_id=str(order.id),
                    notes=ledger_notes,
                    created_by=request.user
                )
                
                # Update qty_received on order item
                order_item.qty_received = order_item.qty_ordered
                order_item.save()
            
            # Update order status to Received
            order.status = 'Received'
            order.save()
        
        # Send email if note is provided
        if note:
            try:
                send_receiving_note_email(order, note, request)
            except Exception as email_error:
                # Log error but don't fail the order processing
                print(f"Error sending receiving note email: {email_error}")
        
        messages.success(request, f'Purchase order {order.po_code} marked as received. All items have been added to inventory.')
        return JsonResponse({
            'success': True,
            'message': f'Purchase order {order.po_code} marked as received successfully!'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def send_receiving_note(request, order_id):
    """Send receiving note email to supplier and order creator (Warehouse Staff only)
    This will cancel the current order and create a new one with the same information"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    
    # Check if user is warehouse staff
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    if not (user_role and 'Warehouse' in user_role):
        return JsonResponse({'success': False, 'error': 'Only warehouse staff can send receiving notes'}, status=403)
    
    try:
        import json
        from django.db import transaction
        from django.utils import timezone
        from datetime import datetime, timedelta
        import secrets
        
        # Get note from request body
        data = json.loads(request.body)
        note = data.get('note', '').strip()
        
        if not note:
            return JsonResponse({'success': False, 'error': 'Note is required'}, status=400)
        
        old_order = get_object_or_404(SupplierOrder, id=order_id)
        
        # Check if order is already cancelled or received
        if old_order.status == 'Cancelled':
            return JsonResponse({'success': False, 'error': 'Order is already cancelled'}, status=400)
        if old_order.status == 'Received':
            return JsonResponse({'success': False, 'error': 'Cannot cancel a received order'}, status=400)
        
        with transaction.atomic():
            # Step 1: Cancel the current order
            old_po_code = old_order.po_code
            old_order.status = 'Cancelled'
            old_order.save()
            
            # Step 2: Create a new order with the same information
            current_year = datetime.now().year
            last_order = SupplierOrder.objects.filter(
                po_code__startswith=f'PO-{current_year}'
            ).order_by('-po_code').first()
            
            if last_order:
                # Extract number from PO code
                # Handle both formats: PO-2026-0001 (old) and PO-20260001 (new)
                po_code_str = last_order.po_code
                if '-' in po_code_str and po_code_str.count('-') == 2:
                    # Old format: PO-2026-0001
                    last_num = int(po_code_str.split('-')[-1])
                else:
                    # New format: PO-20260001 (last 6 digits)
                    last_6_digits = po_code_str[-6:] if len(po_code_str) >= 6 else po_code_str
                    last_num = int(last_6_digits)
                new_num = last_num + 1
            else:
                new_num = 1
            
            new_po_code = f'PO-{current_year}{new_num:06d}'
            
            # Create new supplier order
            new_order = SupplierOrder.objects.create(
                po_code=new_po_code,
                supplier=old_order.supplier,
                created_by=old_order.created_by,
                status='Sent',
                requested_delivery_date=old_order.requested_delivery_date,
                hold_at_supplier=old_order.hold_at_supplier,
            )
            
            # Copy all order items to the new order
            for old_item in old_order.items.all():
                SupplierOrderItem.objects.create(
                    supplier_order=new_order,
                    item=old_item.item,
                    variation=old_item.variation,
                    qty_ordered=old_item.qty_ordered,
                    price_per_unit=old_item.price_per_unit,
                    expected_delivery_date=old_item.expected_delivery_date,
                )
            
            # Generate secure token for new order
            token_string = secrets.token_urlsafe(32)
            expires_at = timezone.now() + timedelta(days=3650)  # 10 years
            
            PortalToken.objects.create(
                token=token_string,
                supplier=new_order.supplier,
                supplier_order=new_order,
                expires_at=expires_at
            )
            
            # Step 3: Send single email with cancellation notice and new invoice link
            try:
                send_receiving_note_email(old_order, new_order, note, request)
                return JsonResponse({
                    'success': True,
                    'message': f'Order cancelled and recreated. Old PO: {old_po_code}, New PO: {new_po_code}'
                })
            except Exception as email_error:
                return JsonResponse({
                    'success': False,
                    'error': f'Error sending email: {str(email_error)}'
                }, status=500)
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def view_purchase_order(request, order_id):
    """View purchase order details (read-only)"""
    order = get_object_or_404(SupplierOrder.objects.select_related('created_by__profile', 'supplier'), id=order_id)
    
    # Check if user is warehouse staff
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_warehouse_staff = user_role and 'Warehouse' in user_role
    
    # Get all order items with details
    order_items = SupplierOrderItem.objects.filter(
        supplier_order=order
    ).select_related('item', 'variation').order_by('id')
    
    # Calculate totals
    items_with_totals = []
    subtotal = Decimal('0.00')
    
    for item in order_items:
        line_total = item.qty_ordered * item.price_per_unit
        subtotal += line_total
        items_with_totals.append({
            'item': item.item,
            'variation': item.variation,
            'quantity': item.qty_ordered,
            'price': item.price_per_unit,
            'line_total': line_total,
            'qty_received': item.qty_received,
        })
    
    total = subtotal
    
    # Get supplier details
    supplier = order.supplier
    
    context = {
        'order': order,
        'supplier': supplier,
        'items': items_with_totals,
        'subtotal': subtotal,
        'total': total,
        'created_by': order.created_by.profile.full_name if hasattr(order.created_by, 'profile') and order.created_by.profile.full_name else (order.created_by.get_full_name() or order.created_by.username),
        'is_warehouse_staff': is_warehouse_staff,
    }
    
    return render(request, 'maainventory/view_purchase_order.html', context)


def view_invoice_by_token(request, token):
    """
    View invoice using secure token (public endpoint, no login required)
    
    Validates token and shows invoice if valid.
    """
    from .models import SupplierInvoiceSignature
    from django.utils import timezone
    from django.http import HttpResponse
    
    # Get token from database
    try:
        portal_token = PortalToken.objects.select_related('supplier_order__created_by__profile', 'supplier_order__supplier').get(token=token)
    except PortalToken.DoesNotExist:
        return render(request, 'maainventory/invoice_error.html', {
            'error_title': 'Invalid Link',
            'error_message': 'This invoice link is invalid or has been removed. Please contact the procurement department for assistance.'
        }, status=404)
    
    # No expiration check - tokens are valid indefinitely
    # (expires_at field kept for backwards compatibility but not enforced)
    
    # Check if token was already used (optional - you can remove this if you want reusable tokens)
    # if portal_token.used_at:
    #     return render(request, 'maainventory/invoice_error.html', {
    #         'error_title': 'Link Already Used',
    #         'error_message': 'This invoice link has already been used. Please contact the procurement department for assistance.'
    #     }, status=403)
    
    order = portal_token.supplier_order
    
    # Check if invoice is already signed
    signature = order.invoice_signatures.first()
    is_signed = signature is not None
    
    # Prepare items with line totals
    items_with_totals = []
    subtotal = 0
    
    for item in order.items.all():
        line_total = item.qty_ordered * item.price_per_unit
        subtotal += line_total
        items_with_totals.append({
            'item': item,
            'line_total': line_total
        })
    
    total = subtotal
    
    context = {
        'order': order,
        'items_with_totals': items_with_totals,
        'subtotal': subtotal,
        'total': total,
        'signature': signature,
        'is_signed': is_signed,
        'portal_token': portal_token,  # Pass token to template for signature linking
    }
    
    return render(request, 'maainventory/invoice.html', context)


def view_invoice(request, order_id):
    """
    View/Download Purchase Order Invoice (by order ID - for internal use)
    
    This view is PUBLICLY ACCESSIBLE (no login required) so suppliers can view
    invoices directly from email links without needing to log in.
    """
    from .models import SupplierInvoiceSignature
    from django.utils import timezone
    
    order = get_object_or_404(SupplierOrder.objects.select_related('created_by__profile', 'supplier'), id=order_id)
    
    # Check if invoice is already signed
    signature = order.invoice_signatures.first()
    is_signed = signature is not None
    
    # Prepare items with line totals
    items_with_totals = []
    subtotal = 0
    
    for item in order.items.all():
        line_total = item.qty_ordered * item.price_per_unit
        subtotal += line_total
        items_with_totals.append({
            'item': item,
            'line_total': line_total
        })
    
    total = subtotal  # Add tax/shipping if needed later
    
    context = {
        'order': order,
        'items_with_totals': items_with_totals,
        'subtotal': subtotal,
        'total': total,
        'signature': signature,
        'is_signed': is_signed,
    }
    
    return render(request, 'maainventory/invoice.html', context)


@csrf_exempt
def submit_invoice_signature(request, order_id=None, token=None):
    """
    Submit invoice signature from supplier (public endpoint, no login required)
    
    Can be accessed via order_id (backward compatibility) or token (secure method).
    
    Note: Using csrf_exempt for public endpoint. In production, consider using
    a token-based approach or signed URLs for better security.
    """
    from .models import SupplierInvoiceSignature, SupplierOrder
    from django.utils import timezone
    from django.http import JsonResponse
    import json
    
    # Get order either by ID or token
    if token:
        try:
            portal_token = PortalToken.objects.select_related('supplier_order__created_by__profile', 'supplier_order__supplier').get(token=token)
            # No expiration check - tokens are valid indefinitely
            order = portal_token.supplier_order
        except PortalToken.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Invalid token'}, status=404)
    else:
        if not order_id:
            return JsonResponse({'success': False, 'error': 'Order ID or token required'}, status=400)
        order = get_object_or_404(SupplierOrder.objects.select_related('created_by__profile', 'supplier'), id=order_id)
        portal_token = order.portal_tokens.first() if hasattr(order, 'portal_tokens') else None
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        # Check if already signed
        if order.invoice_signatures.exists():
            return JsonResponse({'success': False, 'error': 'Invoice already signed'}, status=400)
        
        data = json.loads(request.body)
        signature_data = data.get('signature_data')  # Data URL (data:image/png;base64,...)
        supplier_name = data.get('supplier_name', order.supplier.name)
        
        if not signature_data:
            return JsonResponse({'success': False, 'error': 'Signature data required'}, status=400)
        
        # Extract base64 string from data URL if it includes the prefix
        # Format: data:image/png;base64,iVBORw0KGgo...
        if signature_data.startswith('data:image'):
            # Extract just the base64 part after the comma
            signature_data = signature_data.split(',', 1)[1] if ',' in signature_data else signature_data
        
        # Create signature record
        signature = SupplierInvoiceSignature.objects.create(
            supplier_order=order,
            supplier_name_signed=supplier_name,
            signature_data=signature_data,  # Store only the base64 string
            signed_at=timezone.now(),
            token=portal_token  # Link signature to token if available
        )
        
        # Mark token as used (optional - comment out if you want reusable tokens)
        if portal_token:
            portal_token.used_at = timezone.now()
            portal_token.save()
        
        # Update order status to SIGNED
        order.status = 'Signed'
        order.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Signature submitted successfully',
            'order_id': order.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def reports(request):
    """Comprehensive reports page with all report types"""
    from datetime import datetime, timedelta
    from decimal import Decimal
    from django.db.models import Avg, Min, Max, Count
    from .models import (
        Supplier, SupplierOrder, SupplierOrderItem, SupplierSpendMonthly,
        Item, StockBalance, InventoryLocation, StockLedger,
        Request, RequestItem, Branch, Brand,
        ItemRequest, ItemRequestItem, SupplierStock,
        ItemConsumptionDaily, SupplierCategory, SupplierItem,
        SupplierPriceDiscussion
    )
    
    # Get date range from query params
    # - If start/end are provided (MM/DD/YYYY), use them
    # - Otherwise fall back to "days" (last N days)
    today = datetime.now().date()

    def _parse_date(value: str):
        if not value:
            return None
        value = value.strip()
        try:
            return datetime.strptime(value, "%m/%d/%Y").date()
        except Exception:
            pass
        try:
            # HTML date input format
            return datetime.strptime(value, "%Y-%m-%d").date()
        except Exception:
            return None

    start_str = request.GET.get('start_date')
    end_str = request.GET.get('end_date')
    start_date = _parse_date(start_str)
    end_date = _parse_date(end_str)

    if start_date and end_date:
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        # derive days_back for UI display
        days_back = (end_date - start_date).days if end_date and start_date else 30
    else:
        end_date = today
        try:
            days_back = int(request.GET.get('days', 30))
        except Exception:
            days_back = 30
        start_date = end_date - timedelta(days=days_back)

    created_date_range = (start_date, end_date)
    discussed_date_range = (start_date, end_date)
    consumption_date_range = (start_date, end_date)
    
    # ========================================================================
    # 1. FINANCIAL & SPENDING REPORTS
    # ========================================================================
    
    # Supplier Spending Report
    supplier_spending = []
    suppliers = Supplier.objects.filter(is_active=True).select_related('category')
    
    for supplier in suppliers:
        # Get orders in date range
        orders = SupplierOrder.objects.filter(
            supplier=supplier,
            created_at__date__range=created_date_range
        )
        
        # Calculate total spent
        total_spent = Decimal('0.00')
        order_count = 0
        for order in orders:
            order_items = SupplierOrderItem.objects.filter(supplier_order=order)
            for item in order_items:
                total_spent += item.qty_ordered * item.price_per_unit
            if order_items.exists():
                order_count += 1
        
        if total_spent > 0:
            supplier_spending.append({
                'supplier_name': supplier.name,
                'category': supplier.category.name if supplier.category else 'Uncategorized',
                'total_spent': float(total_spent),
                'order_count': order_count,
                'avg_order_value': float(total_spent / order_count) if order_count > 0 else 0
            })
    
    # Sort by total spent descending
    supplier_spending.sort(key=lambda x: x['total_spent'], reverse=True)
    
    # Purchase Order Financial Summary
    po_summary = {
        'total_orders': SupplierOrder.objects.filter(created_at__date__range=created_date_range).count(),
        'total_value': Decimal('0.00'),
        'by_status': {},
        'avg_order_value': Decimal('0.00')
    }
    
    all_orders = SupplierOrder.objects.filter(created_at__date__range=created_date_range)
    status_counts = {}
    for order in all_orders:
        status = order.get_status_display()
        status_counts[status] = status_counts.get(status, 0) + 1
        
        order_items = SupplierOrderItem.objects.filter(supplier_order=order)
        order_value = sum(item.qty_ordered * item.price_per_unit for item in order_items)
        po_summary['total_value'] += order_value
    
    po_summary['by_status'] = status_counts
    if po_summary['total_orders'] > 0:
        po_summary['avg_order_value'] = po_summary['total_value'] / po_summary['total_orders']
    
    # ========================================================================
    # 2. INVENTORY REPORTS
    # ========================================================================
    
    # Stock Level Report
    stock_levels = []
    warehouse_locations = InventoryLocation.objects.filter(type='WAREHOUSE')
    
    for location in warehouse_locations:
        stock_balances = StockBalance.objects.filter(
            location=location,
            item__is_active=True
        ).select_related('item', 'variation', 'item__brand')
        
        location_total_value = Decimal('0.00')
        low_stock_count = 0
        
        for balance in stock_balances:
            item_value = balance.qty_on_hand * (balance.item.price_per_unit or Decimal('0'))
            location_total_value += item_value
            
            # Check if low stock
            if balance.qty_on_hand < balance.item.min_stock_qty:
                low_stock_count += 1
        
        stock_levels.append({
            'location_name': location.name,
            'total_items': stock_balances.count(),
            'low_stock_count': low_stock_count,
            'total_value': float(location_total_value)
        })
    
    # Low Stock Items
    low_stock_items = []
    all_stock = StockBalance.objects.filter(
        item__is_active=True,
        location__type='WAREHOUSE'
    ).select_related('item', 'variation', 'location')
    
    for stock in all_stock:
        if stock.qty_on_hand < stock.item.min_stock_qty:
            shortage = stock.item.min_stock_qty - stock.qty_on_hand
            low_stock_items.append({
                'item_code': stock.item.item_code,
                'item_name': stock.item.name,
                'variation': stock.variation.variation_name if stock.variation else None,
                'location': stock.location.name,
                'current_stock': float(stock.qty_on_hand),
                'min_stock': float(stock.item.min_stock_qty),
                'shortage': float(shortage),
                'base_unit': stock.item.base_unit
            })
    
    # Stock Movement Report
    stock_movements = StockLedger.objects.filter(
        created_at__date__range=created_date_range
    ).select_related('item', 'variation', 'from_location', 'to_location', 'created_by')[:100]
    
    movement_summary = {
        'total_movements': StockLedger.objects.filter(created_at__date__range=created_date_range).count(),
        'by_reason': {},
        'incoming': Decimal('0.00'),
        'outgoing': Decimal('0.00')
    }
    
    for movement in StockLedger.objects.filter(created_at__date__range=created_date_range):
        reason = movement.get_reason_display()
        movement_summary['by_reason'][reason] = movement_summary['by_reason'].get(reason, 0) + 1
        
        if movement.qty_change > 0:
            movement_summary['incoming'] += movement.qty_change
        else:
            movement_summary['outgoing'] += abs(movement.qty_change)
    
    # ========================================================================
    # 3. OPERATIONAL REPORTS
    # ========================================================================
    
    # Request Performance Report
    requests_data = Request.objects.filter(
        created_at__date__range=created_date_range
    ).select_related('branch', 'branch__brand', 'requested_by', 'approved_by')
    
    request_summary = {
        'total_requests': requests_data.count(),
        'by_status': {},
        'by_branch': {},
        'avg_fulfillment_days': None,
        'approval_rate': None
    }
    
    fulfillment_times = []
    approved_count = 0
    rejected_count = 0
    
    for req in requests_data:
        status = req.get_status_display()
        request_summary['by_status'][status] = request_summary['by_status'].get(status, 0) + 1
        
        branch_name = req.branch.name
        request_summary['by_branch'][branch_name] = request_summary['by_branch'].get(branch_name, 0) + 1
        
        if req.status in ('Approved', 'WarehouseProcessing', 'ReadyForDelivery', 'InProcess', 'OutForDelivery', 'Delivered', 'Completed'):
            approved_count += 1
        elif req.status == 'Rejected':
            rejected_count += 1
        
        # Calculate fulfillment time for completed requests
        if req.status == 'Completed' and req.approved_at:
            days = (req.updated_at.date() - req.approved_at.date()).days
            if days >= 0:
                fulfillment_times.append(days)
    
    if fulfillment_times:
        request_summary['avg_fulfillment_days'] = sum(fulfillment_times) / len(fulfillment_times)
    
    total_reviewed = approved_count + rejected_count
    if total_reviewed > 0:
        request_summary['approval_rate'] = (approved_count / total_reviewed) * 100
    
    # Most Requested Items
    requested_items = RequestItem.objects.filter(
        request__created_at__date__range=created_date_range
    ).values('item__item_code', 'item__name').annotate(
        total_requested=Sum('qty_requested'),
        request_count=Count('request', distinct=True)
    ).order_by('-total_requested')[:20]
    
    # Purchase Order Status Report
    po_status_report = {
        'total_orders': SupplierOrder.objects.filter(created_at__date__range=created_date_range).count(),
        'by_status': {},
        'by_supplier': {},
        'avg_delivery_days': None
    }
    
    delivery_times = []
    for order in SupplierOrder.objects.filter(created_at__date__range=created_date_range).select_related('supplier'):
        status = order.get_status_display()
        po_status_report['by_status'][status] = po_status_report['by_status'].get(status, 0) + 1
        
        supplier_name = order.supplier.name
        po_status_report['by_supplier'][supplier_name] = po_status_report['by_supplier'].get(supplier_name, 0) + 1
    
    # Item Request Report
    item_requests_data = ItemRequest.objects.filter(
        created_at__date__range=created_date_range
    ).select_related('supplier', 'created_by')
    
    item_request_summary = {
        'total_requests': item_requests_data.count(),
        'by_status': {},
        'by_supplier': {},
        'avg_delivery_days_min': None,
        'avg_delivery_days_max': None
    }
    
    delivery_days_min_list = []
    delivery_days_max_list = []
    
    for ir in item_requests_data:
        status = ir.get_status_display()
        item_request_summary['by_status'][status] = item_request_summary['by_status'].get(status, 0) + 1
        
        supplier_name = ir.supplier.name
        item_request_summary['by_supplier'][supplier_name] = item_request_summary['by_supplier'].get(supplier_name, 0) + 1
        
        if ir.delivery_days_min:
            delivery_days_min_list.append(ir.delivery_days_min)
        if ir.delivery_days_max:
            delivery_days_max_list.append(ir.delivery_days_max)
    
    if delivery_days_min_list:
        item_request_summary['avg_delivery_days_min'] = sum(delivery_days_min_list) / len(delivery_days_min_list)
    if delivery_days_max_list:
        item_request_summary['avg_delivery_days_max'] = sum(delivery_days_max_list) / len(delivery_days_max_list)
    
    # ========================================================================
    # 4. ANALYTICS REPORTS
    # ========================================================================
    
    # Item Consumption Report (from Foodics)
    consumption_data = ItemConsumptionDaily.objects.filter(
        date__range=consumption_date_range
    ).select_related('item', 'branch', 'branch__brand', 'variation')
    
    consumption_summary = {
        'total_records': consumption_data.count(),
        'total_consumed': Decimal('0.00'),
        'by_branch': {},
        'by_item': {},
        'top_items': []
    }
    
    for record in consumption_data:
        consumption_summary['total_consumed'] += record.qty_consumed
        
        branch_name = record.branch.name
        consumption_summary['by_branch'][branch_name] = consumption_summary['by_branch'].get(branch_name, Decimal('0')) + record.qty_consumed
        
        item_code = record.item.item_code
        consumption_summary['by_item'][item_code] = consumption_summary['by_item'].get(item_code, Decimal('0')) + record.qty_consumed
    
    # Get top consumed items
    top_consumed = sorted(
        consumption_summary['by_item'].items(),
        key=lambda x: x[1],
        reverse=True
    )[:10]
    
    for item_code, qty in top_consumed:
        item = Item.objects.filter(item_code=item_code).first()
        if item:
            consumption_summary['top_items'].append({
                'item_code': item_code,
                'item_name': item.name,
                'total_consumed': float(qty),
                'base_unit': item.base_unit
            })
    
    # Supplier Performance Report
    supplier_performance = []
    for supplier in suppliers:
        orders = SupplierOrder.objects.filter(supplier=supplier, created_at__date__range=created_date_range)
        
        if orders.exists():
            total_orders = orders.count()
            received_orders = orders.filter(status='Received').count()
            on_time_count = 0  # Would need delivery date tracking for accurate calculation
            
            item_requests = ItemRequest.objects.filter(
                supplier=supplier,
                created_at__date__range=created_date_range
            )
            
            supplier_performance.append({
                'supplier_name': supplier.name,
                'total_orders': total_orders,
                'received_orders': received_orders,
                'completion_rate': (received_orders / total_orders * 100) if total_orders > 0 else 0,
                'item_requests': item_requests.count(),
                'avg_response_days': None  # Would need tracking for accurate calculation
            })
    
    # ========================================================================
    # 5. PRICE DISCUSSION REPORTS
    # ========================================================================
    
    # Latest Price Discussions
    price_discussions_list = []
    price_discussions_qs = SupplierPriceDiscussion.objects.filter(
        discussed_date__date__range=discussed_date_range
    ).select_related('supplier_item', 'supplier_item__supplier', 'supplier_item__item', 'discussed_by').order_by('-discussed_date')[:50]
    
    for discussion in price_discussions_qs:
        # Calculate price difference: from old_price to discussed_price
        if discussion.old_price:
            price_diff = discussion.discussed_price - discussion.old_price
        else:
            # Fallback: compare with current price if old_price wasn't recorded
            price_diff = discussion.discussed_price - discussion.supplier_item.price_per_unit
        
        price_discussions_list.append({
            'discussion': discussion,
            'old_price': float(discussion.old_price) if discussion.old_price else None,
            'new_price': float(discussion.discussed_price),
            'price_difference': float(price_diff)
        })
    
    # Price discussions by supplier
    discussions_by_supplier = {}
    for discussion in SupplierPriceDiscussion.objects.filter(
        discussed_date__date__range=discussed_date_range
    ).select_related('supplier_item__supplier'):
        supplier_name = discussion.supplier_item.supplier.name
        if supplier_name not in discussions_by_supplier:
            discussions_by_supplier[supplier_name] = 0
        discussions_by_supplier[supplier_name] += 1
    
    # Price change trends - get discussions grouped by supplier_item
    price_trends = []
    supplier_items_with_discussions = SupplierItem.objects.filter(
        price_discussions__discussed_date__date__range=discussed_date_range
    ).distinct().select_related('supplier', 'item', 'variation')
    
    for supplier_item in supplier_items_with_discussions:
        discussions = SupplierPriceDiscussion.objects.filter(
            supplier_item=supplier_item,
            discussed_date__date__range=discussed_date_range
        ).order_by('discussed_date')
        
        if discussions.exists():
            # Get price history
            price_history = []
            for disc in discussions:
                price_history.append({
                    'date': disc.discussed_date.date(),
                    'price': float(disc.discussed_price),
                    'discussed_by': disc.discussed_by.profile.full_name if hasattr(disc.discussed_by, 'profile') and disc.discussed_by.profile.full_name else disc.discussed_by.username,
                    'notes': disc.notes or ''
                })
            
            # Calculate price change
            first_price = price_history[0]['price']
            last_price = price_history[-1]['price']
            price_change = last_price - first_price
            price_change_percent = ((price_change / first_price) * 100) if first_price > 0 else 0
            
            # Format price history for JSON serialization
            price_history_json = []
            for disc in discussions:
                price_history_json.append({
                    'date': disc.discussed_date.date().isoformat(),
                    'price': float(disc.discussed_price),
                    'discussed_by': disc.discussed_by.profile.full_name if hasattr(disc.discussed_by, 'profile') and disc.discussed_by.profile.full_name else disc.discussed_by.username,
                    'notes': disc.notes or ''
                })
            
            price_trends.append({
                'supplier_name': supplier_item.supplier.name,
                'item_code': supplier_item.item.item_code,
                'item_name': supplier_item.item.name,
                'variation': supplier_item.variation.variation_name if supplier_item.variation else None,
                'current_price': float(supplier_item.price_per_unit),
                'latest_discussed_price': last_price,
                'first_price': first_price,
                'price_change': price_change,
                'price_change_percent': price_change_percent,
                'discussion_count': discussions.count(),
                'latest_discussion_date': discussions.last().discussed_date.date(),
                'price_history': price_history_json
            })
    
    # Sort by latest discussion date
    price_trends.sort(key=lambda x: x['latest_discussion_date'], reverse=True)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'days_back': days_back,
        # used by <input type="date">
        'start_date_input': start_date.isoformat() if start_date else "",
        'end_date_input': end_date.isoformat() if end_date else "",
        
        # Financial Reports
        'supplier_spending': supplier_spending[:20],  # Top 20
        'po_summary': {
            'total_orders': po_summary['total_orders'],
            'total_value': float(po_summary['total_value']),
            'avg_order_value': float(po_summary['avg_order_value']),
            'by_status': po_summary['by_status']
        },
        
        # Inventory Reports
        'stock_levels': stock_levels,
        'low_stock_items': low_stock_items[:50],  # Top 50
        'stock_movements': list(stock_movements),
        'movement_summary': {
            'total_movements': movement_summary['total_movements'],
            'by_reason': movement_summary['by_reason'],
            'incoming': float(movement_summary['incoming']),
            'outgoing': float(movement_summary['outgoing'])
        },
        
        # Operational Reports
        'request_summary': request_summary,
        'requested_items': list(requested_items),
        'po_status_report': po_status_report,
        'item_request_summary': item_request_summary,
        
        # Analytics Reports
        'consumption_summary': {
            'total_records': consumption_summary['total_records'],
            'total_consumed': float(consumption_summary['total_consumed']),
            'by_branch': {k: float(v) for k, v in consumption_summary['by_branch'].items()},
            'top_items': consumption_summary['top_items']
        },
        'supplier_performance': supplier_performance[:20],  # Top 20
        
        # Price Discussion Reports
        'price_discussions': price_discussions_list,
        'discussions_by_supplier': discussions_by_supplier,
        'price_trends': price_trends[:30],  # Top 30
        'price_trends_json': json.dumps([{
            'label': f"{t['item_code']} - {t['item_name']}",
            'supplier': t['supplier_name'],
            'history': t['price_history']
        } for t in price_trends[:10]])
    }
    
    return render(request, "maainventory/reports.html", context)


@login_required
@csrf_exempt
def add_price_discussion(request):
    """API endpoint to add a price discussion"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method allowed'}, status=405)
    
    try:
        import json
        from django.utils import timezone
        from datetime import datetime
        
        data = json.loads(request.body)
        supplier_item_id = data.get('supplier_item_id')
        discussed_price = data.get('discussed_price')
        discussed_date_str = data.get('discussed_date')
        notes = data.get('notes', '')
        
        if not supplier_item_id or not discussed_price:
            return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)
        
        supplier_item = get_object_or_404(SupplierItem, id=supplier_item_id)
        
        # Parse date
        if discussed_date_str:
            try:
                discussed_date = datetime.fromisoformat(discussed_date_str.replace('Z', '+00:00'))
            except:
                discussed_date = timezone.now()
        else:
            discussed_date = timezone.now()
        
        # Capture the old price BEFORE updating
        old_price = supplier_item.price_per_unit
        new_price = Decimal(str(discussed_price))
        
        # Create price discussion with old price recorded
        discussion = SupplierPriceDiscussion.objects.create(
            supplier_item=supplier_item,
            old_price=old_price,
            discussed_price=new_price,
            discussed_date=discussed_date,
            discussed_by=request.user,
            notes=notes
        )
        
        # Update the supplier item's price to match the discussed price
        supplier_item.price_per_unit = new_price
        supplier_item.save()
        
        # Also update the Item's price_per_unit if it exists
        if supplier_item.item:
            supplier_item.item.price_per_unit = new_price
            supplier_item.item.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Price discussion added and price updated successfully',
            'discussion_id': discussion.id,
            'new_price': str(supplier_item.price_per_unit)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================================
# Branches & Packaging Consumption
# ============================================================================

@login_required
def manage_branch_assignments(request):
    """
    Procurement managers can assign branch managers to branches.
    Lists branch managers and their assigned branches; allows adding/removing assignments.
    """
    from django.http import HttpResponseForbidden
    from .models import BranchUser

    user_profile = getattr(request.user, 'profile', None)
    role_name = (user_profile.role.name if user_profile and user_profile.role else '').lower()
    is_procurement = 'procurement' in role_name
    is_it = 'it' in role_name
    if not (is_procurement or is_it):
        return HttpResponseForbidden('Only Procurement Managers and IT can manage branch assignments.')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            user_id = request.POST.get('user_id')
            branch_ids = request.POST.getlist('branch_ids')
            if user_id:
                if not branch_ids:
                    messages.warning(request, 'Please select at least one branch.')
                else:
                    from django.contrib.auth.models import User
                    user = get_object_or_404(User, id=user_id)
                    added = []
                    for branch_id in branch_ids:
                        branch = Branch.objects.filter(id=branch_id, is_active=True).first()
                        if branch:
                            _, created = BranchUser.objects.get_or_create(user=user, branch=branch)
                            if created:
                                added.append(branch.name)
                    if added:
                        names = ', '.join(added)
                        messages.success(request, f'Assigned {user.get_full_name() or user.username} to: {names}.')
        elif action == 'remove':
            assignment_id = request.POST.get('assignment_id')
            if assignment_id:
                assignment = get_object_or_404(BranchUser, id=assignment_id)
                user_name = assignment.user.get_full_name() or assignment.user.username
                branch_name = assignment.branch.name
                assignment.delete()
                messages.success(request, f'Removed {user_name} from {branch_name}.')
        return redirect('manage_branch_assignments')

    # Get users with Branch Manager role (role name contains "branch" and not "procurement")
    branch_manager_role_ids = list(
        Role.objects.filter(name__icontains='branch')
        .exclude(name__icontains='procurement')
        .values_list('id', flat=True)
    )
    branch_managers = UserProfile.objects.filter(
        role_id__in=branch_manager_role_ids
    ).select_related('user', 'role').order_by('user__username')

    # Get all current assignments for branch managers
    branch_manager_user_ids = [bm.user_id for bm in branch_managers]
    assignments = BranchUser.objects.select_related(
        'user', 'branch', 'branch__brand'
    ).filter(user_id__in=branch_manager_user_ids).order_by('user__username', 'branch__name')

    # Group assignments by user - build list of (branch_manager, assignments) for template
    assignments_by_user = {a.user_id: [] for a in assignments}
    for a in assignments:
        assignments_by_user[a.user_id].append(a)

    manager_rows = []
    for bm in branch_managers:
        manager_rows.append({
            'branch_manager': bm,
            'assignments': assignments_by_user.get(bm.user_id, []),
        })

    # All active branches for the add form, grouped by brand
    all_branches = Branch.objects.filter(is_active=True).select_related('brand').order_by('brand__name', 'name')
    branch_groups = {}
    for branch in all_branches:
        brand_name = branch.brand.name
        if brand_name not in branch_groups:
            branch_groups[brand_name] = []
        branch_groups[brand_name].append(branch)

    context = {
        'branch_managers': branch_managers,
        'manager_rows': manager_rows,
        'all_branches': all_branches,
        'branch_groups': branch_groups,
    }
    return render(request, 'maainventory/manage_branch_assignments.html', context)


@login_required
def branches(request):
    """
    Render branches page: items delivered to each branch; quantity shown is available at branch
    (delivered minus consumption from packaging CSV / other consumption).
    """
    from .context_processors import get_branch_user_info
    from .models import ItemConsumptionDaily

    all_branches = Branch.objects.filter(is_active=True).select_related('brand').order_by('brand__name', 'name')

    is_branch_user, user_branch_ids = get_branch_user_info(request.user)
    if is_branch_user:
        if user_branch_ids:
            all_branches = all_branches.filter(id__in=user_branch_ids)
        else:
            all_branches = all_branches.none()

    # Delivered per branch+item (from fulfilled requests)
    delivered_by_branch_item = {}
    for row in RequestItem.objects.filter(
        request__branch__is_active=True,
        request__status__in=['Delivered', 'Completed'],
        qty_fulfilled__gt=0
    ).values('request__branch_id', 'item_id').annotate(total=Sum('qty_fulfilled')):
        key = (row['request__branch_id'], row['item_id'])
        delivered_by_branch_item[key] = row['total']

    # Consumed per branch+item (packaging CSV, Foodics, etc.)
    consumed_by_branch_item = {}
    for row in ItemConsumptionDaily.objects.filter(branch__is_active=True).values('branch_id', 'item_id').annotate(total=Sum('qty_consumed')):
        key = (row['branch_id'], row['item_id'])
        consumed_by_branch_item[key] = row['total']

    branch_data = []
    for branch in all_branches:
        item_ids = [iid for (bid, iid) in delivered_by_branch_item.keys() if bid == branch.id]
        items_with_qty = []
        if item_ids:
            for item in Item.objects.filter(id__in=item_ids, is_active=True).order_by('item_code'):
                delivered = delivered_by_branch_item.get((branch.id, item.id)) or 0
                consumed = consumed_by_branch_item.get((branch.id, item.id)) or 0
                available = max(Decimal('0'), Decimal(str(delivered)) - Decimal(str(consumed)))
                items_with_qty.append({
                    'item_code': item.item_code,
                    'name': item.name,
                    'base_unit': item.base_unit,
                    'min_order_qty': item.min_order_qty,
                    'qty_delivered': float(delivered),
                    'qty_available': float(available),
                })

        branch_data.append({
            'id': branch.id,
            'name': branch.name,
            'brand': branch.brand.name,
            'address': branch.address or '',
            'items': items_with_qty,
        })

    context = {
        'branch_data': branch_data,
    }
    return render(request, 'maainventory/branches.html', context)


@login_required
def branches_configure(request):
    """Packaging configuration: procurement only. Branch managers do not see this; they use 'Upload CSV to deduct' on the Branches page."""
    from django.http import HttpResponseForbidden
    from .context_processors import get_branch_user_info

    # Only procurement managers can configure packaging rules
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    if not (user_role and 'Procurement' in user_role):
        return HttpResponseForbidden('Only procurement managers can configure packaging rules.')

    all_branches = Branch.objects.filter(is_active=True).select_related('brand').order_by('brand__name', 'name')
    branch_groups = {}
    for branch in all_branches:
        brand_name = branch.brand.name
        if brand_name not in branch_groups:
            branch_groups[brand_name] = []
        branch_groups[brand_name].append({
            'id': branch.id,
            'name': branch.name,
            'address': branch.address or '',
            'rules_count': branch.packaging_rules.count(),
        })
    context = {
        'branch_groups': branch_groups,
    }
    return render(request, 'maainventory/branches_configure.html', context)


@login_required
def branch_packaging(request, branch_id):
    """
    Packaging configuration: Procurement managers can configure which warehouse inventory items
    are used as packaging for each product. Upload CSV to get products, then map to inventory items.
    """
    from django.http import HttpResponseForbidden
    from .context_processors import get_branch_user_info

    # Procurement: define rules only (any branch). Branch managers: define rules + process CSV/deduct for their branch only.
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_procurement = user_role and 'Procurement' in user_role
    is_branch_user, user_branch_ids = get_branch_user_info(request.user)
    
    branch = get_object_or_404(Branch, id=branch_id, is_active=True)
    
    # Branch users can only access their own branch; procurement can access all (for defining rules)
    if not is_procurement:
        if is_branch_user and (not user_branch_ids or branch.id not in user_branch_ids):
            return HttpResponseForbidden('You do not have access to this branch.')
    
    # Only branch managers for this branch can process CSV / deduct; procurement cannot
    can_process_packaging_csv = is_branch_user and user_branch_ids and branch.id in user_branch_ids
    
    rules = branch.packaging_rules.prefetch_related(
        'rule_items__packaging_item',
        'rule_items__inventory_item'
    ).select_related('item').order_by('product_name')
    
    # Items available for this branch = items delivered to this branch (same as /branches/ page)
    branch_item_ids = list(
        RequestItem.objects.filter(
            request__branch_id=branch.id,
            request__status__in=['Delivered', 'Completed'],
            qty_fulfilled__gt=0
        ).values_list('item_id', flat=True).distinct()
    )
    warehouse_items = Item.objects.filter(id__in=branch_item_ids, is_active=True).order_by('name') if branch_item_ids else Item.objects.none()
    
    # Legacy: packaging_items (generic names like Box, Wrapper)
    packaging_items = branch.packaging_items.order_by('display_order', 'name')

    # Check if we have products from a recent upload (define-rules step)
    draft_key = f'branch_packaging_draft_{branch_id}'
    draft_products = request.session.get(draft_key)

    context = {
        'branch': branch,
        'rules': rules,
        'packaging_items': packaging_items,
        'warehouse_items': warehouse_items,
        'draft_products': draft_products,
        'is_procurement': is_procurement,
        'can_process_packaging_csv': can_process_packaging_csv,
    }
    return render(request, 'maainventory/branch_packaging.html', context)


def _parse_products_file(uploaded_file):
    """
    Parse CSV or Excel file to extract product list.
    Required column: Product. Optional: Quantity, Sales, Popularity, Popularity Category.
    Returns list of dicts: [{'product_name': str, 'qty': str, 'sales': str, 'popularity': str, 'popularity_category': str}, ...]
    """
    import csv
    import io
    from openpyxl import load_workbook

    rows = []
    filename = (uploaded_file.name or '').lower()

    if filename.endswith('.csv'):
        content = uploaded_file.read().decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            rows.append(dict(row))
    elif filename.endswith('.xlsx'):
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        # Normalize headers: strip whitespace so "Product", "Quantity" match even with extra spaces/tabs
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        for row in ws.iter_rows(min_row=2):
            vals = [cell.value for cell in row]
            rows.append(dict(zip(headers, vals)))
        wb.close()
    else:
        raise ValueError('File must be CSV or Excel (.csv, .xlsx)')

    def find_col(row, candidates):
        # Match by normalized key (strip, lower) so "Product", "Quantity" work regardless of casing/spaces
        keys = [(k, (k or '').strip().lower()) for k in row.keys() if k]
        for c in candidates:
            cnorm = c.strip().lower()
            for k, knorm in keys:
                if knorm == cnorm:
                    return k
        return None

    result = []
    seen_products = set()
    for row in rows:
        product_col = find_col(row, ['Product', 'product', 'Product Name', 'product_name'])
        if not product_col:
            continue
        product_name = (row.get(product_col) or '').strip()
        if not product_name or product_name in seen_products:
            continue
        seen_products.add(product_name)

        qty_col = find_col(row, ['Quantity', 'quantity', 'Qty', 'qty'])
        sales_col = find_col(row, ['Sales', 'sales'])
        pop_col = find_col(row, ['Popularity', 'popularity'])
        pop_cat_col = find_col(row, ['Popularity Category', 'popularity_category', 'PopularityCategory'])

        result.append({
            'product_name': product_name,
            'qty': str(row.get(qty_col, '') or ''),
            'sales': str(row.get(sales_col, '') or ''),
            'popularity': str(row.get(pop_col, '') or ''),
            'popularity_category': str(row.get(pop_cat_col, '') or ''),
        })
    return result


@login_required
def branch_upload_packaging(request, branch_id):
    """Parse CSV/Excel to extract products, store in session, redirect to define rules form. Procurement managers and branch users can access."""
    from django.http import HttpResponseForbidden
    from .context_processors import get_branch_user_info

    if request.method != 'POST':
        return redirect('branch_packaging', branch_id=branch_id)
    
    # Allow procurement managers to access any branch
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_procurement = user_role and 'Procurement' in user_role
    
    branch = get_object_or_404(Branch, id=branch_id, is_active=True)
    
    if not is_procurement:
        is_branch_user, user_branch_ids = get_branch_user_info(request.user)
        if is_branch_user and (not user_branch_ids or branch.id not in user_branch_ids):
            return HttpResponseForbidden('You do not have access to this branch.')
    uploaded_file = request.FILES.get('packaging_file')
    if not uploaded_file:
        messages.error(request, 'Please select a CSV or Excel file to upload.')
        return redirect('branch_packaging', branch_id=branch_id)
    try:
        products = _parse_products_file(uploaded_file)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect('branch_packaging', branch_id=branch_id)
    except Exception as e:
        messages.error(request, f'Error reading file: {e}')
        return redirect('branch_packaging', branch_id=branch_id)

    if not products:
        messages.warning(request, 'No valid products found in the file. Ensure a "Product" column exists.')
        return redirect('branch_packaging', branch_id=branch_id)

    # Store products in session for the define-rules form
    draft_key = f'branch_packaging_draft_{branch_id}'
    request.session[draft_key] = products
    request.session.modified = True

    messages.success(request, f'Found {len(products)} products. Define packaging rules below (select packaging items per product).')
    return redirect('branch_packaging', branch_id=branch_id)


@login_required
def branch_add_packaging_item(request, branch_id):
    """Add a packaging item (Box, Wrapper, etc.) to the branch. Procurement managers and branch users can access."""
    from django.http import HttpResponseForbidden
    from .context_processors import get_branch_user_info

    if request.method != 'POST':
        return redirect('branch_packaging', branch_id=branch_id)
    
    # Allow procurement managers to access any branch
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_procurement = user_role and 'Procurement' in user_role
    
    branch = get_object_or_404(Branch, id=branch_id, is_active=True)
    
    if not is_procurement:
        is_branch_user, user_branch_ids = get_branch_user_info(request.user)
        if is_branch_user and (not user_branch_ids or branch.id not in user_branch_ids):
            return HttpResponseForbidden('You do not have access to this branch.')
    name = (request.POST.get('packaging_name') or '').strip()
    if not name:
        messages.error(request, 'Please enter a packaging item name.')
        return redirect('branch_packaging', branch_id=branch_id)
    if BranchPackagingItem.objects.filter(branch=branch, name__iexact=name).exists():
        messages.warning(request, f'"{name}" already exists for this branch.')
        return redirect('branch_packaging', branch_id=branch_id)
    max_order = branch.packaging_items.aggregate(m=Max('display_order'))['m'] or 0
    BranchPackagingItem.objects.create(branch=branch, name=name, display_order=max_order + 1)
    messages.success(request, f'Added "{name}" to branch packaging items.')
    return redirect('branch_packaging', branch_id=branch_id)


@login_required
def branch_save_packaging_rules(request, branch_id):
    """
    Save packaging rules from the define-rules form. 
    Maps products to warehouse inventory items (e.g., Mini Kucu Bucket = 1 Burger Box).
    Procurement managers and branch users can configure.
    """
    from django.http import HttpResponseForbidden
    from .context_processors import get_branch_user_info

    if request.method != 'POST':
        return redirect('branch_packaging', branch_id=branch_id)
    
    # Allow procurement managers to access any branch
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_procurement = user_role and 'Procurement' in user_role
    
    branch = get_object_or_404(Branch, id=branch_id, is_active=True)
    
    if not is_procurement:
        is_branch_user, user_branch_ids = get_branch_user_info(request.user)
        if is_branch_user and (not user_branch_ids or branch.id not in user_branch_ids):
            return HttpResponseForbidden('You do not have access to this branch.')

    draft_key = f'branch_packaging_draft_{branch_id}'
    draft_products = request.session.get(draft_key)
    if not draft_products:
        messages.error(request, 'Session expired. Please upload your file again.')
        return redirect('branch_packaging', branch_id=branch_id)

    # Same as branch_packaging: only items delivered to this branch (shown on /branches/)
    branch_item_ids = list(
        RequestItem.objects.filter(
            request__branch_id=branch.id,
            request__status__in=['Delivered', 'Completed'],
            qty_fulfilled__gt=0
        ).values_list('item_id', flat=True).distinct()
    )
    warehouse_items = list(Item.objects.filter(id__in=branch_item_ids, is_active=True).order_by('name')) if branch_item_ids else []
    all_items = {item.name.strip().lower(): item for item in Item.objects.filter(is_active=True)}
    created_count = 0
    updated_count = 0

    for i, p in enumerate(draft_products):
        product_name = p['product_name']
        item = all_items.get(product_name.lower())
        rule, created_flag = BranchPackagingRule.objects.update_or_create(
            branch=branch,
            product_name=product_name,
            defaults={'item': item},
        )
        if created_flag:
            created_count += 1
        else:
            updated_count += 1

        # Clear existing rule items and rebuild from form
        rule.rule_items.all().delete()
        
        # Save rules mapping to warehouse inventory items
        for wh_item in warehouse_items:
            use_key = f'item_use_{i}_{wh_item.id}'
            qty_key = f'item_qty_{i}_{wh_item.id}'
            if request.POST.get(use_key):
                try:
                    qty = Decimal(str(request.POST.get(qty_key, 1) or 1))
                except (ValueError, TypeError):
                    qty = Decimal('1')
                if qty <= 0:
                    qty = Decimal('1')
                BranchPackagingRuleItem.objects.create(
                    rule=rule,
                    inventory_item=wh_item,
                    quantity_per_unit=qty,
                )

    # Clear draft from session
    if draft_key in request.session:
        del request.session[draft_key]
        request.session.modified = True

    messages.success(request, f'Packaging rules saved: {created_count} new, {updated_count} updated.')
    return redirect('branch_packaging', branch_id=branch_id)


@login_required
def branch_cancel_packaging_draft(request, branch_id):
    """Cancel the define-rules step and clear draft from session. Procurement managers and branch users can access."""
    from django.http import HttpResponseForbidden
    from .context_processors import get_branch_user_info

    if request.method != 'POST':
        return redirect('branch_packaging', branch_id=branch_id)
    
    # Allow procurement managers to access any branch
    user_profile = getattr(request.user, 'profile', None)
    user_role = user_profile.role.name if user_profile and user_profile.role else None
    is_procurement = user_role and 'Procurement' in user_role
    
    branch = get_object_or_404(Branch, id=branch_id, is_active=True)
    
    if not is_procurement:
        is_branch_user, user_branch_ids = get_branch_user_info(request.user)
        if is_branch_user and (not user_branch_ids or branch.id not in user_branch_ids):
            return HttpResponseForbidden('You do not have access to this branch.')
    
    draft_key = f'branch_packaging_draft_{branch_id}'
    if draft_key in request.session:
        del request.session[draft_key]
        request.session.modified = True
    messages.info(request, 'Draft cancelled. Upload a new file to define rules.')
    return redirect('branch_packaging', branch_id=branch_id)


@login_required
def branch_process_packaging_csv(request, branch_id):
    """
    Process CSV with Product + Quantity, match to packaging rules, and deduct from this branch's
    available items (items delivered to the branch, shown on /branches/). Does NOT touch warehouse.
    Branch managers only, for their own branch.
    """
    from django.http import HttpResponseForbidden
    from django.db import transaction
    from django.utils import timezone
    from .context_processors import get_branch_user_info
    from .models import ItemConsumptionDaily

    if request.method != 'POST':
        return redirect('branch_packaging', branch_id=branch_id)

    is_branch_user, user_branch_ids = get_branch_user_info(request.user)
    if not is_branch_user or not user_branch_ids or branch_id not in user_branch_ids:
        messages.error(request, 'Only the branch manager for this branch can process packaging CSV.')
        return redirect('branch_packaging', branch_id=branch_id)

    branch = get_object_or_404(Branch, id=branch_id, is_active=True)

    uploaded_file = request.FILES.get('csv_file')
    if not uploaded_file:
        messages.error(request, 'Please select a CSV or Excel file to upload.')
        return redirect('branch_packaging', branch_id=branch_id)

    try:
        products_data = _parse_products_file(uploaded_file)
    except Exception as e:
        messages.error(request, f'Error reading file: {e}')
        return redirect('branch_packaging', branch_id=branch_id)

    if not products_data:
        messages.warning(request, 'No valid products found in the file.')
        return redirect('branch_packaging', branch_id=branch_id)

    # Build deduction plan: which items and how much to deduct at this branch
    deduction_plan = {}
    unmatched_products = []

    for product_data in products_data:
        product_name = product_data.get('product_name', '').strip()
        try:
            product_qty = Decimal(str(product_data.get('qty', 0) or 0))
        except Exception:
            product_qty = Decimal('0')
        if product_qty <= 0:
            continue

        rule = BranchPackagingRule.objects.filter(
            branch=branch,
            product_name__iexact=product_name
        ).prefetch_related('rule_items__inventory_item').first()

        if not rule or not rule.rule_items.exists():
            unmatched_products.append(product_name)
            continue

        for rule_item in rule.rule_items.all():
            if not rule_item.inventory_item:
                continue
            item = rule_item.inventory_item
            qty_needed = product_qty * rule_item.quantity_per_unit
            if item.id not in deduction_plan:
                deduction_plan[item.id] = {'item': item, 'total_qty': Decimal('0')}
            deduction_plan[item.id]['total_qty'] += qty_needed

    if unmatched_products:
        messages.warning(request, f'No packaging rules found for: {", ".join(unmatched_products[:5])}{"..." if len(unmatched_products) > 5 else ""}. Configure rules for these products first.')

    if not deduction_plan:
        messages.warning(request, 'No packaging items to deduct from the uploaded file.')
        return redirect('branch_packaging', branch_id=branch_id)

    # Delivered to this branch (same as /branches/ page)
    delivered = {
        r['item_id']: r['total']
        for r in RequestItem.objects.filter(
            request__branch_id=branch.id,
            request__status__in=['Delivered', 'Completed'],
            qty_fulfilled__gt=0
        ).values('item_id').annotate(total=Sum('qty_fulfilled'))
    }
    # Already consumed at this branch (all sources)
    consumed = {
        r['item_id']: r['total']
        for r in ItemConsumptionDaily.objects.filter(branch=branch).values('item_id').annotate(total=Sum('qty_consumed'))
    }

    insufficient = []
    for item_id, plan in deduction_plan.items():
        item = plan['item']
        qty_needed = plan['total_qty']
        total_delivered = delivered.get(item_id) or Decimal('0')
        total_consumed = consumed.get(item_id) or Decimal('0')
        available = total_delivered - total_consumed
        if available < qty_needed:
            insufficient.append(f'{item.name} (need {qty_needed}, available at branch {available})')

    if insufficient:
        messages.error(request, f'Insufficient quantity at your branch: {"; ".join(insufficient[:3])}{"..." if len(insufficient) > 3 else ""}')
        return redirect('branch_packaging', branch_id=branch_id)

    today = timezone.now().date()
    source = ItemConsumptionDaily.SourceType.PACKAGING_CSV

    try:
        with transaction.atomic():
            for item_id, plan in deduction_plan.items():
                item = plan['item']
                qty_to_deduct = plan['total_qty']
                rec, created = ItemConsumptionDaily.objects.get_or_create(
                    date=today,
                    branch=branch,
                    item=item,
                    variation=None,
                    source=source,
                    defaults={'qty_consumed': qty_to_deduct}
                )
                if not created:
                    rec.qty_consumed += qty_to_deduct
                    rec.save(update_fields=['qty_consumed'])

        messages.success(request, f'Deducted from your branch ({branch.name}): {len(deduction_plan)} item types. Quantities on the Branches page have been updated.')
        return redirect('branch_packaging', branch_id=branch_id)
    except Exception as e:
        import traceback
        traceback.print_exc()
        messages.error(request, f'Error processing packaging: {str(e)}')
        return redirect('branch_packaging', branch_id=branch_id)